"""
form_parser.py — Trích xuất và chuẩn hóa dữ liệu từ file Excel Tờ Đơn (To_Khai_Ung_Vien.xlsx)
Hỗ trợ layout mới với các bảng thân nhân, quá trình học vấn, làm việc ở phần cuối.
"""

import re
from datetime import datetime, date
from io import BytesIO
from typing import Optional
import openpyxl

from core.translator import translate_guardian_name_offline, OFFLINE_JOB_EN, OFFLINE_JOB_JP


def parse_date_str(val) -> Optional[str]:
    """Chuyển đổi các định dạng ngày tháng sang chuỗi chuẩn DD/MM/YYYY hoặc MM/YYYY."""
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.strftime("%d/%m/%Y")
    val_str = str(val).strip()
    if not val_str:
        return None

    # DD/MM/YYYY or DD-MM-YYYY
    m = re.match(r"^(\d{1,2})[/.-](\d{1,2})[/.-](\d{4})$", val_str)
    if m:
        d, mth, y = m.groups()
        return f"{int(d):02d}/{int(mth):02d}/{y}"
    
    # YYYY/MM/DD or YYYY-MM-DD
    m2 = re.match(r"^(\d{4})[/.-](\d{1,2})[/.-](\d{1,2})$", val_str)
    if m2:
        y, mth, d = m2.groups()
        return f"{int(d):02d}/{int(mth):02d}/{y}"

    # MM/YYYY
    m3 = re.match(r"^(\d{1,2})[/.-](\d{4})$", val_str)
    if m3:
        mth, y = m3.groups()
        return f"{int(mth):02d}/{y}"

    return val_str


def _get_val(cell):
    if cell is None or cell.value is None:
        return ""
    val = cell.value
    if isinstance(val, (datetime, date)):
        return val.strftime("%d/%m/%Y")
    if isinstance(val, float):
        if val.is_integer():
            return f"{int(val)}"
        return str(val).strip()
    if isinstance(val, int):
        return str(val).strip()
    return str(val).strip()


def normalize_phone_str(val) -> Optional[str]:
    """Chuẩn hóa chuỗi số điện thoại dạng Text, bảo toàn số 0 đầu nếu Excel lưu dạng số."""
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    if s.endswith(".0"):
        s = s[:-2]
    cleaned = re.sub(r'[\s\.\-]', '', s)
    # Nếu là 9 chữ số bắt đầu bằng các đầu số di động/bàn thông dụng (2,3,5,7,8,9), tự động bù số 0 đầu
    if len(cleaned) == 9 and cleaned[0] in "235789":
        return "0" + cleaned
    # Nếu có tiền tố 84 / +84
    if len(cleaned) == 11 and cleaned.startswith("84") and cleaned[2] in "235789":
        return "0" + cleaned[2:]
    if len(cleaned) == 12 and cleaned.startswith("+84") and cleaned[3] in "235789":
        return "0" + cleaned[3:]
    return s


def normalize_id_number(val, doc_type: str = "CCCD") -> Optional[str]:
    """Chuẩn hóa số giấy tờ CCCD/CMND/Hộ chiếu dạng Text, bảo toàn số 0 đầu."""
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    if s.endswith(".0"):
        s = s[:-2]
    cleaned = re.sub(r'[\s\.\-]', '', s)
    # CCCD chuẩn 12 số: nếu 11 chữ số thì do Excel làm mất số 0 đầu
    if doc_type.upper() == "CCCD" or len(cleaned) == 11:
        if len(cleaned) == 11 and cleaned.isdigit():
            return "0" + cleaned
    # CMND chuẩn 9 số: nếu 8 chữ số thì bù 0 đầu
    if doc_type.upper() == "CMND" or len(cleaned) == 8:
        if len(cleaned) == 8 and cleaned.isdigit():
            return "0" + cleaned
    return s


def parse_candidate_form_excel(file_bytes_or_path) -> dict:
    """Đọc file Excel Tờ Đơn và chuyển đổi sang cấu trúc FullCandidateProfile dictionary."""
    if isinstance(file_bytes_or_path, (bytes, bytearray)):
        wb = openpyxl.load_workbook(BytesIO(file_bytes_or_path), data_only=True)
    else:
        wb = openpyxl.load_workbook(file_bytes_or_path, data_only=True)

    ws = wb.active

    # 1. Candidate Info
    full_name_vn = _get_val(ws["B4"])
    full_name_katakana = _get_val(ws["D4"])
    full_name_en = _get_val(ws["F4"])

    raw_dob = _get_val(ws["B5"])
    date_of_birth = parse_date_str(raw_dob) if raw_dob else None

    gender = _get_val(ws["D5"]) or "Nam"
    phone = normalize_phone_str(_get_val(ws["F5"]))

    marital_status = _get_val(ws["B6"]) or "Độc thân"
    has_children = _get_val(ws["D6"])
    language_skill = _get_val(ws["F6"])

    nationality = _get_val(ws["B7"]) or "Việt Nam"
    ethnicity = _get_val(ws["D7"]) or "Kinh"
    native_language = _get_val(ws["F7"]) or "Tiếng Việt"

    birthplace_vn = _get_val(ws["B8"])
    profile_code = _get_val(ws["F8"])
    address_vn = _get_val(ws["B9"])

    # 2. Identity Docs & Guardian (Rows 11-14)
    cccd_num = normalize_id_number(_get_val(ws["B11"]), "CCCD")
    cccd_date = parse_date_str(_get_val(ws["D11"])) if _get_val(ws["D11"]) else None
    cccd_place = _get_val(ws["F11"])

    passport_num = _get_val(ws["B12"])
    passport_date = parse_date_str(_get_val(ws["D12"])) if _get_val(ws["D12"]) else None
    passport_place = _get_val(ws["F12"])

    guardian_name = _get_val(ws["B13"])
    guardian_rel = _get_val(ws["D13"])
    guardian_phone = normalize_phone_str(_get_val(ws["F13"]))
    guardian_addr = _get_val(ws["B14"])

    # 3. Physical & Health (Rows 16-18)
    def _to_int(val, default=None):
        try:
            return int(float(str(val).replace(",", ".")))
        except (ValueError, TypeError):
            return default

    def _to_float(val, default=None):
        try:
            return float(str(val).replace(",", "."))
        except (ValueError, TypeError):
            return default

    height_cm = _to_float(_get_val(ws["B16"]))
    weight_kg = _to_float(_get_val(ws["D16"]))
    dominant_hand = _get_val(ws["F16"]) or "Phải"
    blood_type = _get_val(ws["H16"]) or "A"

    vision_left = _get_val(ws["B17"])
    vision_right = _get_val(ws["D17"])
    health_status = _get_val(ws["F17"])
    tattoo = _get_val(ws["H17"])

    smoking = _get_val(ws["B18"])
    drinking = _get_val(ws["D18"])
    chronic_disease = _get_val(ws["F18"])
    japan_exp_note = _get_val(ws["H18"])

    # 4. Aspirations & Skills (Rows 20-23)
    internship_field_vn = _get_val(ws["B20"])
    experience_duration = _get_val(ws["F20"])
    japan_goal = _get_val(ws["B21"])
    post_return_plan = _get_val(ws["B22"])
    strengths = _get_val(ws["B23"])
    weaknesses = _get_val(ws["D23"])
    hobbies = _get_val(ws["F23"])

    # Locate Dynamic Tables dynamically or fallback to fixed row indices
    edu_start_row = 26
    work_start_row = 32
    family_start_row = 38
    section_rows = {}

    # Scan for section header markers
    for r in range(1, ws.max_row + 1):
        txt = _get_val(ws.cell(row=r, column=1))
        if "V. QUÁ TRÌNH HỌC VẤN" in txt or "HỌC VẤN" in txt:
            edu_start_row = r + 2
            section_rows["edu"] = r
        elif "VI. QUÁ TRÌNH LÀM VIỆC" in txt or "LÀM VIỆC" in txt:
            work_start_row = r + 2
            section_rows["work"] = r
        elif "VII. THÀNH VIÊN GIA ĐÌNH" in txt or "THÂN NHÂN" in txt:
            family_start_row = r + 2
            section_rows["family"] = r
        elif "VIII." in txt or "cam đoan" in txt.lower():
            section_rows["footer"] = r

    current_year = datetime.now().year

    # 5. Educations (V. Quá trình học vấn)
    educations = []
    edu_end_row = section_rows.get("work", work_start_row - 2)
    for r in range(edu_start_row, edu_end_row):
        start_d = _get_val(ws.cell(row=r, column=2))
        end_d = _get_val(ws.cell(row=r, column=3))
        school = _get_val(ws.cell(row=r, column=4))
        degree = _get_val(ws.cell(row=r, column=7))

        if school or start_d or end_d or degree:
            educations.append({
                "school_name_vn": school or "Trường học",
                "school_name_jp": None,
                "start_date": parse_date_str(start_d) if start_d else None,
                "end_date": parse_date_str(end_d) if end_d else None,
                "education_level": degree or None,
                "degree_level_vn": degree or None,
            })

    # 6. Work Experiences (VI. Quá trình làm việc)
    work_experiences = []
    work_end_row = section_rows.get("family", family_start_row - 2)
    for r in range(work_start_row, work_end_row):
        start_d = _get_val(ws.cell(row=r, column=2))
        end_d = _get_val(ws.cell(row=r, column=3))
        company = _get_val(ws.cell(row=r, column=4))
        job = _get_val(ws.cell(row=r, column=6))

        if company or start_d or end_d or job:
            work_experiences.append({
                "company_name_vn": company or "Công ty",
                "company_name_jp": None,
                "start_date": parse_date_str(start_d) if start_d else None,
                "end_date": parse_date_str(end_d) if end_d else None,
                "job_title_vn": job or None,
                "job_title_jp": None,
                "job_description_vn": job or None,
                "description": job or None,
            })

    # 7. Family Members (VII. Thân nhân gia đình)
    family_members = []
    family_end_row = section_rows.get("footer", ws.max_row + 1)
    for r in range(family_start_row, family_end_row):
        rel = _get_val(ws.cell(row=r, column=2))
        name = _get_val(ws.cell(row=r, column=3))
        byear_raw = _get_val(ws.cell(row=r, column=5))
        job = _get_val(ws.cell(row=r, column=6))
        income = _get_val(ws.cell(row=r, column=7)) if ws.max_column >= 7 else None
        cohab = _get_val(ws.cell(row=r, column=8))

        if name or rel or byear_raw or job:
            byear_num = _to_int(byear_raw)
            calc_age = None
            if byear_num:
                if byear_num > 1900 and byear_num <= current_year:
                    calc_age = current_year - byear_num
                elif byear_num < 150:
                    calc_age = byear_num

            is_cohab = (cohab.lower() in ("có", "co", "yes", "true", "1", "o", "⭕") if cohab else True)
            job_en = OFFLINE_JOB_EN.get(job.lower(), None) if job else None
            job_jp = OFFLINE_JOB_JP.get(job.lower(), None) if job else None

            family_members.append({
                "relationship": rel or "Người thân",
                "relationship_vn": rel or "Người thân",
                "full_name": name or "Thành viên",
                "age": calc_age,
                "birth_year": byear_num,
                "living_together": "Có" if is_cohab else "Không",
                "is_living_together": is_cohab,
                "occupation": job or None,
                "occupation_vn": job or None,
                "occupation_en": job_en,
                "occupation_jp": job_jp,
                "workplace": None,
                "monthly_income": income or None,
            })

    # Identity Documents
    identity_docs = []
    if cccd_num:
        identity_docs.append({
            "document_type": "CCCD",
            "doc_type": "CCCD",
            "document_number": cccd_num,
            "issue_date": cccd_date,
            "issue_place_vn": cccd_place or "Cục Cảnh sát QLHC về TTXH",
            "issue_place": cccd_place or "Cục Cảnh sát QLHC về TTXH",
            "is_primary": True,
        })
    if passport_num:
        identity_docs.append({
            "document_type": "Passport",
            "doc_type": "Passport",
            "document_number": passport_num,
            "issue_date": passport_date,
            "issue_place_vn": passport_place or "Cục Quản lý xuất nhập cảnh",
            "issue_place": passport_place or "Cục Quản lý xuất nhập cảnh",
            "is_primary": False,
        })

    # Find guardian's job from family members if available
    guardian_job_vn = None
    guardian_job_en = None
    guardian_job_jp = None
    if guardian_rel:
        for fm in family_members:
            if fm.get("relationship") and fm.get("relationship").lower() in (guardian_rel.lower(), f"{guardian_rel.lower()} đẻ"):
                guardian_job_vn = fm.get("occupation")
                guardian_job_en = fm.get("occupation_en")
                guardian_job_jp = fm.get("occupation_jp")
                break

    guardian_name_en_str = None
    if guardian_name:
        raw_g_str = f"{guardian_name} ({guardian_rel})" if guardian_rel else guardian_name
        guardian_name_en_str = translate_guardian_name_offline(raw_g_str)

    candidate_dict = {
        "full_name_vn": full_name_vn,
        "full_name_katakana": full_name_katakana or None,
        "full_name_eng": full_name_en or None,
        "date_of_birth": date_of_birth,
        "gender": gender,
        "phone": phone or None,
        "marital_status": marital_status,
        "has_children": has_children or "Không",
        "foreign_languages": language_skill or None,
        "nationality": nationality,
        "ethnicity": ethnicity,
        "mother_tongue": native_language,
        "birthplace_vn": birthplace_vn or None,
        "address_vn": address_vn or None,
        "guardian_name_vn": guardian_name or None,
        "guardian_name": guardian_name or None,
        "guardian_name_en": guardian_name_en_str,
        "guardian_relationship": guardian_rel or None,
        "guardian_job_vn": guardian_job_vn,
        "guardian_job_en": guardian_job_en,
        "guardian_job_jp": guardian_job_jp,
        "guardian_address_vn": guardian_addr or None,
        "guardian_address": guardian_addr or None,
        "guardian_phone": guardian_phone or None,
        "height_cm": height_cm,
        "weight_kg": weight_kg,
        "blood_type": blood_type,
        "vision_left": vision_left or None,
        "vision_right": vision_right or None,
        "preferred_hand": dominant_hand,
        "dominant_hand": dominant_hand,
        "tattoos": tattoo or "Không",
        "tattoo": tattoo or "Không",
        "smoking": smoking or "Không",
        "alcohol": drinking or "Không",
        "health_status": health_status or "Tốt",
        "chronic_disease": chronic_disease or "Không",
        "purpose_to_japan_vn": japan_goal or None,
        "japan_goal_vn": japan_goal or None,
        "plan_after_return_vn": post_return_plan or None,
        "strengths_vn": strengths or None,
        "weaknesses_vn": weaknesses or None,
        "hobbies_vn": hobbies or None,
        "hobby_vn": hobbies or None,
        "profile_code": profile_code or None,
    }

    assignment_dict = {
        "internship_field_vn": internship_field_vn or None,
    }

    return {
        "candidate": candidate_dict,
        "identityDocuments": identity_docs,
        "educations": educations,
        "workExperiences": work_experiences,
        "familyMembers": family_members,
        "assignment": assignment_dict,
    }

