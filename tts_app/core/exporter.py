"""
exporter.py — Export candidates to File_lưu.xlsx (60-col format)
"""
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from datetime import datetime
import re

from core.models import Candidate

# 60-column headers for the master export sheet
COL60_HEADERS = [
    "STT", "MA HO SO", "TEN VNM", "TEN ENG", "TEN PHIEN AM", "GIOI TINH JPN",
    "SO CAN CUOC", "NGAY CAP CAN CUOC VNM", "NGAY CAP CAN CUOC JPN",
    "NOI CAP CAN CUOC VNM", "NOI CAP CAN CUOC JPN",
    "SO HO CHIEU", "NGAY CAP HO CHIEU VNM", "NGAY CAP HO CHIEU JPN",
    "NOI CAP HO CHIEU VNM", "NOI CAP HO CHIEU JPN",
    "NAM SINH VNM", "NAM SINH JPN",
    "TINH TRANG HON NHAN", "CO CON",
    "DIA CHI VNM", "DIA CHI JPN", "NOI SINH VNM", "NOI SINH JPN",
    "NGUOI GIAM HO VNM", "NGUOI GIAM HO JPN",
    "DIA CHI NGUOI GIAM HO VNM", "DIA CHI NGUOI GIAM HO JPN",
    "SDT NGUOI GIAM HO",
    "TRUONG HOC 1 VNM", "TRUONG HOC 1 JPN",
    "TRUONG HOC 2 VNM", "TRUONG HOC 2 JPN",
    "TRUONG HOC 3 VNM", "TRUONG HOC 3 JPN",
    "CONG TY 1 VNM", "CONG TY 1 JPN",
    "CONG TY 2 VNM", "CONG TY 2 JPN",
    "CONG TY 3 VNM", "CONG TY 3 JPN",
    "LINH VUC THUC TAP VNM", "LINH VUC THUC TAP JPN",
    "TOM TAT KN JPN", "TOM TAT KN VNM",
    "NGHIEP DOAN VNM", "NGHIEP DOAN JPN",
    "DC NGHIEP DOAN VNM", "DC NGHIEP DOAN JPN",
    "ND NGHIEP DOAN VNM", "ND NGHIEP DOAN JPN",
    "CONG TY TIEP NHAN VNM", "CONG TY TIEP NHAN JPN",
    "DC CONG TY TIEP NHAN VNM", "DC CONG TY TIEP NHAN JPN",
    "ND CONG TY TIEP NHAN VNM", "ND CONG TY TIEP NHAN JPN",
    "CONG TY PHAI CU VNM", "ND CONG TY PHAI CU VNM",
    "SO DIEN THOAI",
]


# ── Style helpers ─────────────────────────────────────────────────────────────

_HEADER_FILL  = PatternFill("solid", fgColor="1E3A5F")
_HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=9)
_DATA_FONT    = Font(name="Arial", size=9)
_CENTER       = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT         = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_THIN         = Side(style="thin", color="AAAAAA")
_BORDER       = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_FILL_ALT     = PatternFill("solid", fgColor="F0F4FA")

# Column widths (approximate)
_COL_WIDTHS = [
    5, 12, 22, 22, 22, 8, 18,
    18, 18, 30, 30,
    16, 18, 18, 30, 30,
    18, 18, 8, 10,
    40, 40, 20, 20,
    35, 35, 40, 40, 16,
    20, 35, 20, 35, 20, 35,
    20, 35, 20, 35, 20, 35,
    25, 25, 12, 12,
    35, 35, 40, 40, 25, 25,
    35, 35, 40, 40, 25, 25,
    30, 25, 16,
]


def _style_header_row(ws, row: int, ncols: int):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill   = _HEADER_FILL
        cell.font   = _HEADER_FONT
        cell.border = _BORDER
        cell.alignment = _CENTER


def _style_data_row(ws, row: int, ncols: int, alt: bool = False):
    fill = _FILL_ALT if alt else None
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        if fill:
            cell.fill = fill
        cell.font   = _DATA_FONT
        cell.border = _BORDER
        cell.alignment = _LEFT


def _set_col_widths(ws, widths: list):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── Main sheet builder ────────────────────────────────────────────────────────

COL60_KEYS = [
    "id", "profile_code", "full_name_vn", "full_name_eng", "full_name_katakana", "gender",
    "id_document_number", "id_issue_date", "id_issue_date_jp", "id_issue_place_vn", "id_issue_place_jp",
    "passport_number", "passport_issue_date", "passport_issue_date_jp", "passport_issue_place_vn", "passport_issue_place_jp",
    "date_of_birth", "date_of_birth_jp", "marital_status", "has_children",
    "address_vn", "address_jp", "birthplace_vn", "birthplace_jp",
    "guardian_name_vn", "guardian_name_jp", "guardian_address_vn", "guardian_address_jp", "guardian_phone",
    "school1_period", "school1_name", "school2_period", "school2_name", "school3_period", "school3_name",
    "work1_period", "work1_name", "work2_period", "work2_name", "work3_period", "work3_name",
    "internship_field_vn", "internship_field_jp", "skill_summary_jp", "skill_summary_vn",
    "syndicate_name_vn", "syndicate_name_jp", "syndicate_address_vn", "syndicate_address_jp", "syndicate_rep_vn", "syndicate_rep_jp",
    "company_name_vn", "company_name_jp", "company_address_vn", "company_address_jp", "company_rep_vn", "company_rep_jp",
    "dispatching_company_vn", "dispatching_company_rep_vn", "phone"
]


def build_master_sheet(ws, candidates: list, sheet_name: str = None):
    """Write candidates to a master sheet with 60-col format."""
    ws.title = sheet_name or f"Thang {datetime.now().strftime('%m')}"
    ws.freeze_panes = "B2"

    # Header
    for col_idx, h in enumerate(COL60_HEADERS, 1):
        ws.cell(row=1, column=col_idx, value=h)
    _style_header_row(ws, 1, len(COL60_HEADERS))
    _set_col_widths(ws, _COL_WIDTHS)
    ws.row_dimensions[1].height = 30

    text_keys = {
        "profile_code", "id_document_number", "id_issue_date", "id_issue_date_jp", 
        "passport_number", "passport_issue_date", "passport_issue_date_jp", 
        "date_of_birth", "date_of_birth_jp", "guardian_phone", "phone"
    }

    for row_idx, cand in enumerate(candidates, 2):
        alt = (row_idx % 2 == 0)
        for col_idx, key in enumerate(COL60_KEYS, 1):
            val = cand.get(key) if isinstance(cand, dict) else getattr(cand, key, None)
            if key == "id":
                val = row_idx - 1  # STT
            if val is None:
                val = ""
            if key in text_keys and val != "":
                val = str(val)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if key in text_keys:
                cell.number_format = "@"
        _style_data_row(ws, row_idx, len(COL60_KEYS), alt)
        ws.row_dimensions[row_idx].height = 20

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COL60_HEADERS))}1"


def build_syndicate_sheet(ws, syndicates: list):
    ws.title = "Nghiệp đoàn"
    headers = [
        "STT", "TEN ND VNM", "TEN ND JPN",
        "CHU TICH ND VNM", "CHU TICH ND JPN",
        "D/C NGHIEP DOAN VNM", "D/C NGHIEP DOAN JPN", "SDT JAPAN"
    ]
    fields = [
        "id", "ten_vnm", "ten_jpn",
        "chu_tich_vnm", "chu_tich_jpn",
        "dia_chi_vnm", "dia_chi_jpn", "so_dien_thoai"
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=2, column=c, value=h)
    _style_header_row(ws, 2, len(headers))
    ws.column_dimensions["A"].width = 5
    for c in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 35

    for r, s in enumerate(syndicates, 3):
        for c, f in enumerate(fields, 1):
            val = s.get(f) if isinstance(s, dict) else getattr(s, f, None)
            if f == "id":
                val = r - 2
            cell = ws.cell(row=r, column=c, value=str(val) if val is not None else "")
            if f == "so_dien_thoai":
                cell.number_format = "@"
        _style_data_row(ws, r, len(fields), r % 2 == 0)


def build_company_sheet(ws, companies: list):
    ws.title = "Chủ sử dụng"
    headers = [
        "STT", "TEN TO CHUC THUC TAP VNM", "TEN TO CHUC THUC TAP JPN",
        "TEN GIAM DOC VNM", "TEN GD JPN",
        "D/C THUC TAP VNM", "D/C THUC TAP JPN", "SO DIEN THOAI"
    ]
    fields = [
        "id", "ten_vnm", "ten_jpn",
        "giam_doc_vnm", "giam_doc_jpn",
        "dia_chi_vnm", "dia_chi_jpn", "so_dien_thoai"
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=2, column=c, value=h)
    _style_header_row(ws, 2, len(headers))
    ws.column_dimensions["A"].width = 5
    for c in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 35

    for r, co in enumerate(companies, 3):
        for c, f in enumerate(fields, 1):
            val = co.get(f) if isinstance(co, dict) else getattr(co, f, None)
            if f == "id":
                val = r - 2
            cell = ws.cell(row=r, column=c, value=str(val) if val is not None else "")
            if f == "so_dien_thoai":
                cell.number_format = "@"
        _style_data_row(ws, r, len(fields), r % 2 == 0)


def export_to_excel(
    candidates: list,
    syndicates: list,
    companies: list,
    output_path: str,
    sheet_name: str = None,
):
    """
    Build and save File_lưu.xlsx with:
      - Sheet 1: Master 60 cột
      - Sheet 2: Nghiệp đoàn
      - Sheet 3: Chủ sử dụng
    """
    wb = openpyxl.Workbook()

    # Sheet 1 — Master
    ws1 = wb.active
    build_master_sheet(ws1, candidates, sheet_name)

    # Sheet 2 — Syndicates
    ws2 = wb.create_sheet()
    build_syndicate_sheet(ws2, syndicates)

    # Sheet 3 — Companies
    ws3 = wb.create_sheet()
    build_company_sheet(ws3, companies)

    wb.save(output_path)
    return output_path
