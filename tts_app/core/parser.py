"""
parser.py — Parse CVpv.xlsx (履歴書 Nhật Bản format) → candidate dict (60 cột)

Layout cố định mỗi sheet:
  Row 1  (idx 0): 番号：XX → mã số ứng viên
  Row 4  (idx 3): col 4 = Furigana (Katakana)
  Row 5  (idx 4): col 4 = Latin name (TEN ENG)
  Row 7  (idx 6): col 4 = DOB JP, col 12 = age text, col 17 = gender text
  Row 9  (idx 8): col 4 = Nơi sinh JP
  Row 10 (idx 9): col 4 = Địa chỉ JP
  Row 11 (idx10): col 4 = Hôn nhân, col 23 = Dân tộc
  Row 13 (idx12): education row 1 (period col4, school col10)
  Row 14 (idx13): education row 2
  Row 15 (idx14): education row 3
  Row 17 (idx16): work row 1 (period col4, company col10, job col20)
  Row 18 (idx17): work row 2
  Row 19 (idx18): work row 3
  Row 21 (idx20): col 4 = ngoại ngữ / tiếng Nhật
  Row 26 (idx25): cha (続柄 col0, name col2, age col10, job col14)
  Row 27 (idx26): mẹ
  Row 28 (idx27): anh/chị/em/vợ/chồng
  Row 35 (idx34): chiều cao col3, tay thuận col6
  Row 36 (idx35): cân nặng col3
  Row 37 (idx36): nhóm máu col6 (blank in test data — inferred from col mapping)
  Row 40 (idx39): col 9 = mục đích
  Row 41 (idx40): col 9 = sau thực tập
  Row 42 (idx41): col 3 = ưu điểm, col 12 = nhược điểm, col 23 = sở thích
"""

import re
import openpyxl
from datetime import datetime
from typing import Optional


def _cell(rows: list, row_idx: int, col_idx: int) -> Optional[str]:
    """Safe cell getter."""
    try:
        v = rows[row_idx][col_idx]
        return str(v).strip() if v is not None else None
    except (IndexError, TypeError):
        return None


def _clean(s) -> Optional[str]:
    if s is None:
        return None
    s = str(s).strip()
    # Remove zero-width spaces and special chars
    s = re.sub(r'[\u200b\u200e\u200f\ufeff]', '', s)
    # Normalize whitespace inside Japanese strings (keep ideographic space)
    return s if s else None


def _extract_age(text: str) -> Optional[str]:
    """Extract age from '年齢 (26)歳' → '26歳'"""
    if not text:
        return None
    m = re.search(r'\((\d+)\)', text)
    if m:
        return f"{m.group(1)}歳"
    return None


def _extract_age_vnm(text: str) -> Optional[str]:
    """Extract age from '年齢 (26)歳' → '26 tuổi'"""
    if not text:
        return None
    m = re.search(r'\((\d+)\)', text)
    if m:
        return f"{m.group(1)} tuổi"
    return None


def _infer_gender(row: tuple) -> Optional[str]:
    """From gender row, try to determine 男/女.
    The row typically has '男　・　女' as a merged cell — we check adjacent cells for marks."""
    # Cell at col 17 often has gender text
    # We just return None here; user fills or GSheet provides
    return None


def _parse_dob_vnm(dob_jpn: str) -> Optional[str]:
    """Convert '2000年10月28日' → '2000-10-28'"""
    if not dob_jpn:
        return None
    m = re.match(r'(\d{4})年(\d{1,2})月(\d{1,2})日', dob_jpn)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    return dob_jpn


def _scan_education(rows: list) -> list:
    """
    Scan rows 12-15 (idx 11-14) for education entries.
    Returns list of (period, school_name) tuples.
    """
    entries = []
    for idx in range(11, 16):
        period = _clean(_cell(rows, idx, 4))
        school = _clean(_cell(rows, idx, 10))
        if period and ('年' in period or '～' in period):
            entries.append((period, school or ""))
    return entries


def _scan_work(rows: list) -> list:
    """
    Scan rows 16-20 (idx 15-20) for work history.
    Returns list of (period, company_job) tuples.
    """
    entries = []
    for idx in range(15, 21):
        period  = _clean(_cell(rows, idx, 4))
        company = _clean(_cell(rows, idx, 10))
        job     = _clean(_cell(rows, idx, 20))
        if period and ('年' in period or '～' in period or '現在' in period):
            if company and job:
                label = f"{company}（{job}）"
            elif company:
                label = company
            elif job:
                label = job
            else:
                label = ""
            entries.append((period, label))
    return entries


def _scan_family(rows: list) -> list:
    """
    Scan rows 25-30 (idx 24-30) for family members.
    Returns list of dicts with relationship, name, age, job.
    """
    family = []
    relations_jp = {'父', '母', '兄', '姉', '弟', '妹', '妻', '夫', '子', '祖父', '祖母'}
    for idx in range(24, 31):
        rel  = _clean(_cell(rows, idx, 0))
        name = _clean(_cell(rows, idx, 2))
        age  = _clean(_cell(rows, idx, 10))
        job  = _clean(_cell(rows, idx, 14))
        if rel and name and (rel in relations_jp or len(rel) <= 3):
            family.append({
                "rel_jp":  rel,
                "name":    name,
                "age":     age,
                "job":     job,
            })
    return family


def _detect_guardian(family: list) -> dict:
    """Choose guardian from family (prefer 父, then 母)."""
    for member in family:
        if member["rel_jp"] == "父":
            return member
    for member in family:
        if member["rel_jp"] == "母":
            return member
    if family:
        return family[0]
    return {}


_REL_MAP = {
    "父": "Cha", "母": "Mẹ", "兄": "Anh", "姉": "Chị",
    "弟": "Em trai", "妹": "Em gái", "妻": "Vợ", "夫": "Chồng",
    "子": "Con", "祖父": "Ông", "祖母": "Bà",
}


def parse_cv_sheet(ws) -> dict:
    """Parse a single 履歴書 worksheet → candidate dict."""
    rows = list(ws.iter_rows(values_only=True))

    # ── Row 1: candidate number ───────────────────────────────────────────────
    num_cell = _clean(_cell(rows, 0, 0)) or ""
    m = re.search(r'(\d+)', num_cell)
    candidate_num = int(m.group(1)) if m else 0

    # ── Row 4: Katakana name ──────────────────────────────────────────────────
    ten_phien_am = _clean(_cell(rows, 3, 4))

    # ── Row 5: Latin name ─────────────────────────────────────────────────────
    ten_eng = _clean(_cell(rows, 4, 4))

    # ── Row 7: DOB, age, gender ───────────────────────────────────────────────
    dob_jpn  = _clean(_cell(rows, 6, 4))
    dob_vnm  = _parse_dob_vnm(dob_jpn)
    age_text = _clean(_cell(rows, 6, 12))
    tuoi_jpn = _extract_age(age_text)
    tuoi_vnm = _extract_age_vnm(age_text)
    gender_cell = _clean(_cell(rows, 6, 17))
    gioi_tinh = None  # Need manual input or GSheet

    # ── Row 9: Nơi sinh ────────────────────────────────────────────────────────
    noi_sinh_jpn = _clean(_cell(rows, 8, 4))

    # ── Row 10: Địa chỉ ───────────────────────────────────────────────────────
    dia_chi_jpn = _clean(_cell(rows, 9, 4))

    # ── Row 11: Hôn nhân, Dân tộc ────────────────────────────────────────────
    hon_nhan = _clean(_cell(rows, 10, 4))
    dan_toc  = _clean(_cell(rows, 10, 23))

    # ── Education ─────────────────────────────────────────────────────────────
    edu = _scan_education(rows)
    qua_trinh_hoc = [None, None, None]
    ten_truong    = [None, None, None]
    for i, (p, s) in enumerate(edu[:3]):
        qua_trinh_hoc[i] = p
        ten_truong[i]    = s

    # ── Work history ──────────────────────────────────────────────────────────
    work = _scan_work(rows)
    qt_lam_viec = [None, None, None]
    ten_dn      = [None, None, None]
    for i, (p, c) in enumerate(work[:3]):
        qt_lam_viec[i] = p
        ten_dn[i]      = c

    # ── Row 21: Tiếng Nhật ────────────────────────────────────────────────────
    tieng_nhat_cell = _clean(_cell(rows, 20, 4))

    # ── Row 21: Người thân ở Nhật ─────────────────────────────────────────────
    nguoi_than_nhat = _clean(_cell(rows, 30, 3))

    # ── Family ────────────────────────────────────────────────────────────────
    family  = _scan_family(rows)
    guardian = _detect_guardian(family)
    if guardian:
        rel_vn = _REL_MAP.get(guardian.get("rel_jp", ""), guardian.get("rel_jp", ""))
        guardian_name_vnm = f"{guardian['name']} ({rel_vn})"
        guardian_name_jpn = f"{guardian['name']} （{guardian.get('rel_jp','')}）"
        guardian_job      = guardian.get("job", "")
    else:
        guardian_name_vnm = guardian_name_jpn = guardian_job = None

    # ── Physical ──────────────────────────────────────────────────────────────
    chieu_cao = _clean(_cell(rows, 34, 3))
    tay_thuan = _clean(_cell(rows, 35, 6))
    can_nang  = _clean(_cell(rows, 35, 3))
    # Blood type is at row 37 col 6 but often empty in test data
    nhom_mau  = _clean(_cell(rows, 36, 6))
    if not nhom_mau:
        nhom_mau = _clean(_cell(rows, 36, 3))

    # ── Motivation / Plans / Traits ───────────────────────────────────────────
    muc_dich    = _clean(_cell(rows, 39, 9))
    sau_tap     = _clean(_cell(rows, 40, 9))
    uu_diem     = _clean(_cell(rows, 41, 3))
    nhuoc_diem  = _clean(_cell(rows, 41, 12))
    so_thich    = _clean(_cell(rows, 41, 23))

    # Auto generate mã hồ sơ placeholder (user can edit)
    ma_ho_so = f"TTS-{candidate_num:03d}"

    return {
        # Identity
        "ma_ho_so":      ma_ho_so,
        "ten_vnm":       ten_eng,        # VNM name = same as Latin initially; user adds diacritics
        "ten_eng":       ten_eng,
        "ten_phien_am":  ten_phien_am,
        "gioi_tinh_jpn": gioi_tinh,
        "sdt_tts":       None,

        # CCCD (not in CV — must be filled)
        "so_can_cuoc":       None,
        "ngay_cap_cccd_vnm": None,
        "ngay_cap_cccd_jpn": None,
        "noi_cap_cccd_vnm":  None,
        "noi_cap_cccd_jpn":  None,

        # Passport (not in CV)
        "so_ho_chieu":     None,
        "ngay_cap_hc_vnm": None,
        "ngay_cap_hc_jpn": None,
        "noi_cap_hc_vnm":  None,
        "noi_cap_hc_jpn":  None,

        # DOB & Age
        "nam_sinh_vnm": dob_vnm,
        "nam_sinh_jpn": dob_jpn,
        "tuoi_jpn":     tuoi_jpn,
        "tuoi_vnm":     tuoi_vnm,

        # Address
        "dia_chi_vnm":  None,
        "dia_chi_jpn":  dia_chi_jpn,
        "noi_sinh_vnm": None,
        "noi_sinh_jpn": noi_sinh_jpn,

        # Guardian
        "nguoi_giam_ho_vnm": guardian_name_vnm,
        "nguoi_giam_ho_jpn": guardian_name_jpn,
        "dc_nguoi_gh_vnm":   None,
        "dc_nguoi_gh_jpn":   None,
        "sdt_nguoi_gh":      None,

        # Education
        "qua_trinh_hoc_1": qua_trinh_hoc[0],
        "ten_truong_1":    ten_truong[0],
        "qua_trinh_hoc_2": qua_trinh_hoc[1],
        "ten_truong_2":    ten_truong[1],
        "qua_trinh_hoc_3": qua_trinh_hoc[2],
        "ten_truong_3":    ten_truong[2],

        # Work
        "qt_lam_viec_1": qt_lam_viec[0],
        "ten_dn_1":       ten_dn[0],
        "qt_lam_viec_2": qt_lam_viec[1],
        "ten_dn_2":       ten_dn[1],
        "qt_lam_viec_3": qt_lam_viec[2],
        "ten_dn_3":       ten_dn[2],

        # Job info (blank from CV)
        "nganh_nghe_vnm":  None,
        "nganh_nghe_jpn":  None,
        "kinh_nghiem_jpn": None,
        "kinh_nghiem_vnm": None,

        # Syndicate & Company (blank from CV)
        "ten_nghiep_doan_vnm":  None,
        "ten_nghiep_doan_jpn":  None,
        "dc_nghiep_doan_vnm":   None,
        "dc_nghiep_doan_jpn":   None,
        "ten_chu_tich_nd_vnm":  None,
        "ten_chu_tich_nd_jpn":  None,
        "ten_tiep_nhan_vnm":    None,
        "ten_tiep_nhan_jpn":    None,
        "dc_tiep_nhan_vnm":     None,
        "dc_tiep_nhan_jpn":     None,
        "ten_gd_tiep_nhan_vnm": None,
        "ten_gd_tiep_nhan_jpn": None,
        "cong_ty_chung_nghe":   None,
        "ten_gd_chung_nghe":    None,

        # Extra CV fields
        "hon_nhan":        hon_nhan,
        "dan_toc":         dan_toc,
        "chieu_cao":       chieu_cao,
        "can_nang":        can_nang,
        "nhom_mau":        nhom_mau,
        "tay_thuan":       tay_thuan,
        "suc_khoe":        None,
        "trinh_do_nhat":   tieng_nhat_cell,
        "nguoi_than_nhat": nguoi_than_nhat,
        "muc_dich":        muc_dich,
        "sau_thuc_tap":    sau_tap,
        "uu_diem":         uu_diem,
        "nhuoc_diem":      nhuoc_diem,
        "so_thich":        so_thich,

        "status": "Mới tạo",
    }


def parse_cv_file(filepath: str) -> list:
    """
    Parse all sheets in CVpv.xlsx.
    Returns list of candidate dicts.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    results = []
    errors  = []
    for sh_name in wb.sheetnames:
        ws = wb[sh_name]
        # Skip non-CV sheets
        first_cell = str(ws.cell(1, 1).value or "").strip()
        if not first_cell.startswith("番号"):
            continue
        try:
            data = parse_cv_sheet(ws)
            data["_sheet"] = sh_name.strip()
            results.append(data)
        except Exception as e:
            errors.append({"sheet": sh_name, "error": str(e)})
    return results, errors
