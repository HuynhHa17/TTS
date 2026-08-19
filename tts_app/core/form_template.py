"""
form_template.py — Tạo file Excel Mẫu Tờ Đơn / Phiếu Đăng Ký Thông Tin Ứng Viên TTS
Layout mới:
  I. Thông tin cá nhân
  II. Giấy tờ tùy thân & Người giám hộ
  III. Thể lực & Sức khỏe
  IV. Nguyện vọng thực tập & Năng khiếu
  --- Các mục phát sinh dòng để ở cuối ---
  V. Quá trình học vấn
  VI. Quá trình làm việc
  VII. Thành viên gia đình (Thân nhân)
  VIII. Cam kết & Ký tên
"""

import os
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def create_candidate_form_workbook() -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "To_Khai_Ung_Vien"

    # ── Setup Page Setup for Standard A4 (1 Page Single Sheet) ──
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_A4  # A4 standard (210 x 297 mm)
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1

    # Margins (inches) - Compact & balanced for A4
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.4
    ws.page_margins.header = 0.2
    ws.page_margins.footer = 0.2

    # Centered horizontally on A4 page
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = False

    # Show grid lines
    ws.views.sheetView[0].showGridLines = True

    # Print Area
    ws.print_area = "A1:H46"

    # ── Style Definitions ──
    font_title = Font(name="Times New Roman", size=13, bold=True, color="FFFFFF")
    font_sub_title = Font(name="Times New Roman", size=8.5, italic=True, color="DDE2E5")
    font_sec_hdr = Font(name="Times New Roman", size=9.5, bold=True, color="FFFFFF")
    font_lbl = Font(name="Times New Roman", size=8.5, bold=True, color="1A1A1A")
    font_inp = Font(name="Times New Roman", size=9.5, bold=False, color="002060")
    font_guide = Font(name="Times New Roman", size=8, italic=True, color="888888")

    fill_title = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    fill_sec_hdr = PatternFill(start_color="2B4C7E", end_color="2B4C7E", fill_type="solid")
    fill_lbl = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")
    fill_inp = PatternFill(start_color="FFFDF0", end_color="FFFDF0", fill_type="solid")  # light warm tint for input
    fill_photo = PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid")

    thin_border_side = Side(border_style="thin", color="B0BEC5")
    border_cell = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Column widths optimized for A4 printable width (Total ~110 units)
    col_widths = {
        "A": 14,  # Nhãn 1
        "B": 18,  # Giá trị 1
        "C": 12,  # Nhãn 2
        "D": 15,  # Giá trị 2
        "E": 12,  # Nhãn 3
        "F": 16,  # Giá trị 3
        "G": 12,  # Nhãn 4
        "H": 15,  # Giá trị 4
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    def style_range(cell_range, fill=None, font=None, alignment=None, border=border_cell, number_format=None):
        for row in ws[cell_range]:
            for cell in row:
                if fill: cell.fill = fill
                if font: cell.font = font
                if alignment: cell.alignment = alignment
                if border: cell.border = border
                if number_format: cell.number_format = number_format

    def set_cell(coord, val, fill=fill_inp, font=font_inp, alignment=align_left, number_format="@", **kwargs):
        c = ws[coord]
        c.value = val
        if fill: c.fill = fill
        if font: c.font = font
        if alignment: c.alignment = alignment
        if number_format: c.number_format = number_format
        c.border = border_cell

    def set_label(coord, val, fill=fill_lbl, font=font_lbl, alignment=align_left, **kwargs):
        c = ws[coord]
        c.value = val
        if fill: c.fill = fill
        if font: c.font = font
        if alignment: c.alignment = alignment
        c.border = border_cell

    # ══════════════════════════════════════════════════════════════════════════
    # ROW 1-2: HEADER BANNER
    # ══════════════════════════════════════════════════════════════════════════
    ws.merge_cells("A1:H1")
    ws["A1"] = "PHIẾU ĐĂNG KÝ THÔNG TIN THỰC TẬP SINH NHẬT BẢN"
    style_range("A1:H1", fill=fill_title, font=font_title, alignment=align_center)
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    ws["A2"] = "Vui lòng điền đầy đủ và chính xác các thông tin dưới đây vào các ô có màu nền vàng nhạt"
    style_range("A2:H2", fill=fill_title, font=font_sub_title, alignment=align_center)
    ws.row_dimensions[2].height = 18

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION I: THÔNG TIN CÁ NHÂN (Rows 3 - 9)
    # ══════════════════════════════════════════════════════════════════════════
    ws.merge_cells("A3:H3")
    ws["A3"] = "I. THÔNG TIN CÁ NHÂN"
    style_range("A3:H3", fill=fill_sec_hdr, font=font_sec_hdr, alignment=align_left)
    ws.row_dimensions[3].height = 22

    # Row 4: Họ tên VN, Katakana, Ảnh 4x6
    set_label("A4", "Họ và tên (VN) *")
    set_cell("B4", "")
    set_label("C4", "Phiên âm Katakana")
    set_cell("D4", "")
    set_label("E4", "Tên tiếng Anh")
    set_cell("F4", "")
    
    # Ô dán ảnh 4x6 (G4:H7)
    ws.merge_cells("G4:H7")
    ws["G4"] = "ẢNH 4x6\n(Dán ảnh tại đây)"
    style_range("G4:H7", fill=fill_photo, font=font_guide, alignment=align_center)

    # Row 5: Ngày sinh, Giới tính, SĐT
    set_label("A5", "Ngày sinh (DD/MM/YYYY) *")
    set_cell("B5", "")
    set_label("C5", "Giới tính *")
    set_cell("D5", "Nam")
    set_label("E5", "Số điện thoại *")
    set_cell("F5", "")

    # Row 6: Hôn nhân, Có con, Ngoại ngữ
    set_label("A6", "Tình trạng hôn nhân")
    set_cell("B6", "Độc thân")
    set_label("C6", "Có con chưa?")
    set_cell("D6", "Không")
    set_label("E6", "Trình độ ngoại ngữ")
    set_cell("F6", "Tiếng Nhật sơ cấp")

    # Row 7: Quốc tịch, Dân tộc, Tôn giáo
    set_label("A7", "Quốc tịch")
    set_cell("B7", "Việt Nam")
    set_label("C7", "Dân tộc")
    set_cell("D7", "Kinh")
    set_label("E7", "Tiếng mẹ đẻ")
    set_cell("F7", "Tiếng Việt")

    # Row 8: Nơi sinh & Mã hồ sơ
    set_label("A8", "Nơi sinh (Tỉnh/TP) *")
    ws.merge_cells("B8:D8")
    style_range("B8:D8", fill=fill_inp, font=font_inp, alignment=align_left, number_format="@")
    ws["B8"] = ""

    set_label("E8", "Mã hồ sơ (Nếu có)")
    ws.merge_cells("F8:H8")
    style_range("F8:H8", fill=fill_inp, font=font_inp, alignment=align_left, number_format="@")
    ws["F8"] = ""

    # Row 9: Địa chỉ thường trú
    set_label("A9", "Địa chỉ thường trú *")
    ws.merge_cells("B9:H9")
    style_range("B9:H9", fill=fill_inp, font=font_inp, alignment=align_left, number_format="@")
    ws["B9"] = ""

    for r in range(4, 10):
        ws.row_dimensions[r].height = 22

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION II: GIẤY TỜ TÙY THÂN & NGƯỜI GIÁM HỘ (Rows 10 - 14)
    # ══════════════════════════════════════════════════════════════════════════
    ws.merge_cells("A10:H10")
    ws["A10"] = "II. GIẤY TỜ TÙY THÂN & NGƯỜI GIÁM HỘ"
    style_range("A10:H10", fill=fill_sec_hdr, font=font_sec_hdr, alignment=align_left)
    ws.row_dimensions[10].height = 22

    # CCCD (Row 11)
    set_label("A11", "Số CCCD / CMND *")
    set_cell("B11", "", number_format="@")
    set_label("C11", "Ngày cấp CCCD")
    set_cell("D11", "", number_format="@")
    set_label("E11", "Nơi cấp CCCD")
    ws.merge_cells("F11:H11")
    style_range("F11:H11", fill=fill_inp, font=font_inp, alignment=align_left, number_format="@")
    ws["F11"] = "Cục Cảnh sát QLHC về TTXH"

    # Passport (Row 12)
    set_label("A12", "Số Hộ chiếu (Passport)")
    set_cell("B12", "", number_format="@")
    set_label("C12", "Ngày cấp Hộ chiếu")
    set_cell("D12", "", number_format="@")
    set_label("E12", "Nơi cấp Hộ chiếu")
    ws.merge_cells("F12:H12")
    style_range("F12:H12", fill=fill_inp, font=font_inp, alignment=align_left, number_format="@")
    ws["F12"] = "Cục Quản lý xuất nhập cảnh"

    # Người giám hộ (Row 13 - 14)
    set_label("A13", "Tên người giám hộ *")
    set_cell("B13", "")
    set_label("C13", "Quan hệ (Bố/Mẹ/...)")
    set_cell("D13", "Bố")
    set_label("E13", "Số điện thoại GH *")
    ws.merge_cells("F13:H13")
    style_range("F13:H13", fill=fill_inp, font=font_inp, alignment=align_left, number_format="@")
    ws["F13"] = ""

    set_label("A14", "Địa chỉ người giám hộ")
    ws.merge_cells("B14:H14")
    style_range("B14:H14", fill=fill_inp, font=font_inp, alignment=align_left, number_format="@")
    ws["B14"] = ""

    for r in range(11, 15):
        ws.row_dimensions[r].height = 22

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION III: THỂ LỰC & SỨC KHỎE (Rows 15 - 18)
    # ══════════════════════════════════════════════════════════════════════════
    ws.merge_cells("A15:H15")
    ws["A15"] = "III. THỂ LỰC & TÌNH TRẠNG SỨC KHỎE"
    style_range("A15:H15", fill=fill_sec_hdr, font=font_sec_hdr, alignment=align_left)
    ws.row_dimensions[15].height = 22

    set_label("A16", "Chiều cao (cm)")
    set_cell("B16", "168")
    set_label("C16", "Cân nặng (kg)")
    set_cell("D16", "60")
    set_label("E16", "Tay thuận")
    set_cell("F16", "Phải")
    set_label("G16", "Nhóm máu")
    set_cell("H16", "A")

    set_label("A17", "Thị lực mắt trái")
    set_cell("B17", "10/10")
    set_label("C17", "Thị lực mắt phải")
    set_cell("D17", "10/10")
    set_label("E17", "Tình trạng sức khỏe")
    set_cell("F17", "Tốt")
    set_label("G17", "Hình xăm")
    set_cell("H17", "Không")

    set_label("A18", "Hút thuốc?")
    set_cell("B18", "Không")
    set_label("C18", "Uống rượu bia?")
    set_cell("D18", "Không")
    set_label("E18", "Bệnh mãn tính")
    set_cell("F18", "Không")
    set_label("G18", "Từng đi Nhật?")
    set_cell("H18", "Không")

    for r in range(16, 19):
        ws.row_dimensions[r].height = 22

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION IV: NGUYỆN VỌNG THỰC TẬP & NĂNG KHIẾU (Rows 19 - 23)
    # ══════════════════════════════════════════════════════════════════════════
    ws.merge_cells("A19:H19")
    ws["A19"] = "IV. NGUYỆN VỌNG THỰC TẬP & NĂNG KHIẾU"
    style_range("A19:H19", fill=fill_sec_hdr, font=font_sec_hdr, alignment=align_left)
    ws.row_dimensions[19].height = 22

    set_label("A20", "Ngành nghề đăng ký")
    ws.merge_cells("B20:D20")
    style_range("B20:D20", fill=fill_inp, font=font_inp, alignment=align_left, number_format="@")
    ws["B20"] = "Thi công máy móc xây dựng"

    set_label("E20", "Thời gian kinh nghiệm")
    ws.merge_cells("F20:H20")
    style_range("F20:H20", fill=fill_inp, font=font_inp, alignment=align_left, number_format="@")
    ws["F20"] = "2 năm"

    set_label("A21", "Mục đích sang Nhật")
    ws.merge_cells("B21:H21")
    style_range("B21:H21", fill=fill_inp, font=font_inp, alignment=align_left, number_format="@")
    ws["B21"] = "Học hỏi kỹ thuật tiên tiến và tích lũy vốn phát triển tương lai"

    set_label("A22", "Kế hoạch sau về nước")
    ws.merge_cells("B22:H22")
    style_range("B22:H22", fill=fill_inp, font=font_inp, alignment=align_left, number_format="@")
    ws["B22"] = "Làm việc tại doanh nghiệp Nhật Bản tại Việt Nam"

    set_label("A23", "Ưu điểm")
    set_cell("B23", "Chăm chỉ, nhanh nhẹn")
    set_label("C23", "Nhược điểm")
    set_cell("D23", "Ít nói")
    set_label("E23", "Sở thích")
    ws.merge_cells("F23:H23")
    style_range("F23:H23", fill=fill_inp, font=font_inp, alignment=align_left, number_format="@")
    ws["F23"] = "Thể thao, đọc sách"

    for r in range(20, 24):
        ws.row_dimensions[r].height = 22

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION V: QUÁ TRÌNH HỌC VẤN (Rows 24 - 29)
    # ══════════════════════════════════════════════════════════════════════════
    ws.merge_cells("A24:H24")
    ws["A24"] = "V. QUÁ TRÌNH HỌC VẤN (Từ cấp 2 đến nay)"
    style_range("A24:H24", fill=fill_sec_hdr, font=font_sec_hdr, alignment=align_left)
    ws.row_dimensions[24].height = 22

    set_label("A25", "STT", alignment=align_center)
    set_label("B25", "Từ (MM/YYYY)", alignment=align_center)
    set_label("C25", "Đến (MM/YYYY)", alignment=align_center)
    ws.merge_cells("D25:F25")
    style_range("D25:F25", fill=fill_lbl, font=font_lbl, alignment=align_center)
    ws["D25"] = "Tên trường học"
    ws.merge_cells("G25:H25")
    style_range("G25:H25", fill=fill_lbl, font=font_lbl, alignment=align_center)
    ws["G25"] = "Trình độ / Bằng cấp (THCS/THPT/ĐH...)"
    ws.row_dimensions[25].height = 22

    # Edu Rows (26 - 29) -> 4 rows
    default_edus = ["THCS", "THPT", "Trung cấp / Cao đẳng", "Đại học"]
    for i, r in enumerate(range(26, 30), start=1):
        ws.row_dimensions[r].height = 21
        set_cell(f"A{r}", str(i), alignment=align_center)
        set_cell(f"B{r}", "")
        set_cell(f"C{r}", "")
        ws.merge_cells(f"D{r}:F{r}")
        style_range(f"D{r}:F{r}", fill=fill_inp, font=font_inp, alignment=align_left, number_format="@")
        ws.merge_cells(f"G{r}:H{r}")
        style_range(f"G{r}:H{r}", fill=fill_inp, font=font_inp, alignment=align_left, number_format="@")
        ws[f"G{r}"] = default_edus[i-1] if i <= len(default_edus) else ""

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION VI: QUÁ TRÌNH LÀM VIỆC (Rows 30 - 35)
    # ══════════════════════════════════════════════════════════════════════════
    ws.merge_cells("A30:H30")
    ws["A30"] = "VI. QUÁ TRÌNH LÀM VIỆC (Kinh nghiệm thực tế)"
    style_range("A30:H30", fill=fill_sec_hdr, font=font_sec_hdr, alignment=align_left)
    ws.row_dimensions[30].height = 22

    set_label("A31", "STT", alignment=align_center)
    set_label("B31", "Từ (MM/YYYY)", alignment=align_center)
    set_label("C31", "Đến (MM/YYYY)", alignment=align_center)
    ws.merge_cells("D31:E31")
    style_range("D31:E31", fill=fill_lbl, font=font_lbl, alignment=align_center)
    ws["D31"] = "Tên công ty / Doanh nghiệp"
    ws.merge_cells("F31:H31")
    style_range("F31:H31", fill=fill_lbl, font=font_lbl, alignment=align_center)
    ws["F31"] = "Vị trí / Nội dung công việc cụ thể"
    ws.row_dimensions[31].height = 22

    # Work Rows (32 - 35) -> 4 rows
    for i, r in enumerate(range(32, 36), start=1):
        ws.row_dimensions[r].height = 21
        set_cell(f"A{r}", str(i), alignment=align_center)
        set_cell(f"B{r}", "")
        set_cell(f"C{r}", "")
        ws.merge_cells(f"D{r}:E{r}")
        style_range(f"D{r}:E{r}", fill=fill_inp, font=font_inp, alignment=align_left, number_format="@")
        ws.merge_cells(f"F{r}:H{r}")
        style_range(f"F{r}:H{r}", fill=fill_inp, font=font_inp, alignment=align_left, number_format="@")

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION VII: THÀNH VIÊN GIA ĐÌNH / THÂN NHÂN (Rows 36 - 42)
    # ══════════════════════════════════════════════════════════════════════════
    ws.merge_cells("A36:H36")
    ws["A36"] = "VII. THÀNH VIÊN GIA ĐÌNH (THÂN NHÂN)"
    style_range("A36:H36", fill=fill_sec_hdr, font=font_sec_hdr, alignment=align_left)
    ws.row_dimensions[36].height = 22

    set_label("A37", "STT", alignment=align_center)
    set_label("B37", "Quan hệ", alignment=align_center)
    ws.merge_cells("C37:D37")
    style_range("C37:D37", fill=fill_lbl, font=font_lbl, alignment=align_center)
    ws["C37"] = "Họ và tên thành viên"
    set_label("E37", "Tuổi / Năm sinh", alignment=align_center)
    set_label("F37", "Nghề nghiệp", alignment=align_center)
    set_label("G37", "Thu nhập/tháng", alignment=align_center)
    set_label("H37", "Sống chung?", alignment=align_center)
    ws.row_dimensions[37].height = 22

    # Family Rows (38 - 42) -> 5 rows
    default_rels = ["Bố", "Mẹ", "Anh/Em", "Chị/Em", "Vợ/Chồng"]
    for i, r in enumerate(range(38, 43), start=1):
        ws.row_dimensions[r].height = 21
        set_cell(f"A{r}", str(i), alignment=align_center)
        set_cell(f"B{r}", default_rels[i-1] if i <= len(default_rels) else "")
        ws.merge_cells(f"C{r}:D{r}")
        style_range(f"C{r}:D{r}", fill=fill_inp, font=font_inp, alignment=align_left, number_format="@")
        set_cell(f"E{r}", "")
        set_cell(f"F{r}", "")
        set_cell(f"G{r}", "")
        set_cell(f"H{r}", "Có", alignment=align_center)

    # ══════════════════════════════════════════════════════════════════════════
    # FOOTER & CAM KẾT (Rows 43 - 46)
    # ══════════════════════════════════════════════════════════════════════════
    ws.merge_cells("A43:H43")
    ws["A43"] = "Tôi xin cam đoan những thông tin khai trên đây là hoàn toàn đúng sự thật."
    style_range("A43:H43", font=Font(name="Times New Roman", size=9, italic=True, bold=True), alignment=align_center)
    ws.row_dimensions[43].height = 20

    ws.merge_cells("F44:H44")
    ws["F44"] = "Ngày ...... tháng ...... năm 202..."
    style_range("F44:H44", font=Font(name="Times New Roman", size=9, italic=True), alignment=align_center)

    ws.merge_cells("F45:H45")
    ws["F45"] = "Người làm đơn (Ký và ghi rõ họ tên)"
    style_range("F45:H45", font=Font(name="Times New Roman", size=9, bold=True), alignment=align_center)

    ws.row_dimensions[46].height = 40  # signature space

    return wb


def export_candidate_form_template(output_path: str = None) -> bytes:
    wb = create_candidate_form_workbook()
    if output_path:
        wb.save(output_path)
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()
