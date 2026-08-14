"""
API: Excel Import/Export — đọc & ghi File_lưu.xlsx (60 cột)
Endpoints:
  GET  /api/excel/config           — lấy đường dẫn file hiện tại
  POST /api/excel/config           — đổi đường dẫn file Excel
  POST /api/excel/import           — import từ Excel → SQLite (SQLite là master)
  GET  /api/excel/export           — xuất SQLite → ghi vào Excel
  GET  /api/excel/preview          — preview dữ liệu Excel (không import)
"""
import os
import re
from datetime import datetime, date
import json
from io import BytesIO

from flask import Blueprint, request, jsonify, current_app, send_file
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from core.database import get_session
from core.models import (
    Candidate, IdentityDocument, Education, WorkExperience,
    SkillExperience, FamilyMember, CandidateAssignment, Organization,
    AppSettings, to_dict
)
import config

excel_io_bp = Blueprint("excel_io", __name__)

# ─────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────

def _get_excel_path(db) -> str:
    """Lấy đường dẫn file Excel đã cấu hình (mặc định = config.OUTPUT_FILE)."""
    row = db.query(AppSettings).filter(AppSettings.key == "excel_output_path").first()
    return row.value if row and row.value else config.OUTPUT_FILE


def _str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return str(v)
    return str(v).strip()


def _to_jp_date(vn_str: str) -> str:
    """Chuyển '28/10/2000' hoặc '2000-10-28' → '2000年10月28日'."""
    if not vn_str:
        return ""
    try:
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d"):
            try:
                d = datetime.strptime(vn_str.strip(), fmt)
                return f"{d.year}年{d.month:02d}月{d.day:02d}日"
            except ValueError:
                pass
    except Exception:
        pass
    return vn_str


def _parse_date_str(val) -> str:
    """Chuyển openpyxl date/datetime → chuỗi YYYY-MM-DD."""
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        return val.strftime("%Y-%m-%d")
    return _str(val)


# ─────────────────────────────────────────────────────────────
# HEADER ORDER của File_lưu.xlsx (60 cột, cột 1 = STT auto)
# ─────────────────────────────────────────────────────────────
HEADERS = [
    "STT",                          # 1  — tự đánh số
    "MA HO SO",                     # 2
    "TEN VNM",                      # 3
    "TEN ENG",                      # 4
    "TEN PHIEN AM",                 # 5
    "GIOI TINH JPN",                # 6
    "SO CAN CUOC",                  # 7
    "NGAY CAP CAN CUOC VNM",        # 8
    "NGAY CAP CAN CUOC JPN",        # 9
    "NOI CAP CAN CUOC VNM",         # 10
    "NOI CAP CAN CUOC JPN",         # 11
    "SO HO CHIEU",                  # 12
    "NGAY CAP HO CHIEU VNM",        # 13
    "NGAY CAP HO CHIEU JPN",        # 14
    "NOI CAP HO CHIEU VNM",         # 15
    "NOI CAP HO CHIEU JPN",         # 16
    "NAM SINH VNM",                 # 17
    "NAM SINH JPN",                 # 18
    "TUOI JPN",                     # 19
    "TUOI VNM",                     # 20
    "DIA CHI VNM",                  # 21
    "DIA CHI JPN",                  # 22
    "NOI SINH VNM",                 # 23
    "NOI SINH JPN",                 # 24
    "NGUOI GIAM HO, QUAN HE (VNM)", # 25
    "NGUOI GIAM HO, QUAN HE (JPN)", # 26
    "DIA CHI NGUOI GIAM HO (VNM)",  # 27
    "DIA CHI NGUOI GIAM HO (JPN)",  # 28
    "SDT NGUOI GIAM HO",            # 29
    "QUA TRINH HOC 1",              # 30
    "TEN TRUONG 1",                 # 31
    "QUA TRINH HOC 2",              # 32
    "TEN TRUONG 2",                 # 33
    "QUA TRINH HOC 3",              # 34
    "TEN TRUONG  3",                # 35
    "QT LAM VIEC 1",                # 36
    "TEN DOANH NGHIEP 1 (NGANH NGHE)", # 37
    "QT LAM VIEC 2",                # 38
    "TEN DOANH NGHIEP 2 ( NGANH NGHE)", # 39
    "QT LAM VIEC 3",                # 40
    "TEN DOANH NGHIEP 3 ( NGANH NGHE)", # 41
    "NGANH NGHE TTS VMN",           # 42
    "NGANH NGHE TTS JPN",           # 43
    "KINH NGHIEM JPN",              # 44
    "KINH NGHIEM VNM",              # 45
    "TEN NGHIEP DOAN VNM",          # 46
    "TEN NGHIEP DOAN JPN",          # 47
    "D/C NGHIEP DOAN VNM",          # 48
    "D/C NGHIEP DOAN JPN",          # 49
    "TEN CHU  TICH ND VNM",         # 50
    "TEN CHU TICH ND JPN",          # 51
    "TEN TIEP NHAN VNM",            # 52
    "TEN TIEP NHAN JPN",            # 53
    "D/C TIEP NHAN VNM",            # 54
    "D/C TIEP NHAN JPN",            # 55
    "TEN GD TIEP NHAN VNM",         # 56
    "TEN GD TIEP NHAN JPN",         # 57
    "CONG TY CHUNG NGHE",           # 58
    "TEN GIAM DOC CTY CHUNG NGHE",  # 59
    "SỐ ĐT TTS",                    # 60
]


def _candidate_to_row(c: Candidate, stt: int) -> list:
    """Chuyển 1 Candidate (+ relations đã load) → list 60 giá trị."""
    # ── Giấy tờ
    cccd = next((d for d in c.identity_documents if d.document_type == "CCCD"), None)
    passport = next((d for d in c.identity_documents if d.document_type == "Passport"), None)

    # ── Học vấn (tối đa 3)
    edus = list(c.educations)[:3]
    while len(edus) < 3:
        edus.append(None)

    # ── Công việc (tối đa 3)
    works = list(c.work_experiences)[:3]
    while len(works) < 3:
        works.append(None)

    # ── Assignment & Orgs
    asgn = c.assignment
    sup_org = asgn.supervising_org if asgn else None
    acc_org = asgn.accepting_org if asgn else None
    send_org = asgn.sending_org if asgn else None

    # ── Tuổi
    age_str = ""
    if c.date_of_birth:
        try:
            for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
                try:
                    dob = datetime.strptime(c.date_of_birth.strip(), fmt)
                    age = (datetime.now() - dob).days // 365
                    age_str = str(age)
                    break
                except ValueError:
                    pass
        except Exception:
            pass

    dob_jp = c.date_of_birth_jp or _to_jp_date(c.date_of_birth)

    def _edu_period(e):
        if not e:
            return ""
        parts = []
        if e.start_date:
            parts.append(e.start_date)
        if e.end_date:
            parts.append(e.end_date)
        return "   ～ ".join(parts) if parts else ""

    def _work_period(w):
        if not w:
            return ""
        parts = []
        if w.start_date:
            parts.append(w.start_date)
        if w.end_date:
            parts.append(w.end_date)
        return "   ～ ".join(parts) if parts else ""

    def _work_name(w):
        if not w:
            return ""
        parts = [_str(w.company_name_vn)]
        if w.job_title_vn:
            parts.append(f"({w.job_title_vn})")
        return " ".join(p for p in parts if p)

    return [
        stt,                                                    # 1
        _str(c.profile_code),                                   # 2
        _str(c.full_name_vn),                                   # 3
        _str(c.full_name_eng),                                   # 4
        _str(c.full_name_katakana),                             # 5
        _str(c.gender),                                         # 6
        _str(cccd.document_number if cccd else ""),             # 7
        _str(cccd.issue_date if cccd else ""),                  # 8
        _str(cccd.issue_date_jp if cccd else ""),               # 9
        _str(cccd.issue_place_vn if cccd else ""),              # 10
        _str(cccd.issue_place_jp if cccd else ""),              # 11
        _str(passport.document_number if passport else ""),     # 12
        _str(passport.issue_date if passport else ""),          # 13
        _str(passport.issue_date_jp if passport else ""),       # 14
        _str(passport.issue_place_vn if passport else ""),      # 15
        _str(passport.issue_place_jp if passport else ""),      # 16
        _str(c.date_of_birth),                                  # 17
        dob_jp,                                                 # 18
        f"{age_str}歳" if age_str else "",                      # 19
        f"{age_str} tuổi" if age_str else "",                   # 20
        _str(c.address_vn),                                     # 21
        _str(c.address_jp),                                     # 22
        _str(c.birthplace_vn),                                  # 23
        _str(c.birthplace_jp),                                  # 24
        _str(c.guardian_name_vn),                               # 25
        _str(c.guardian_name_jp),                               # 26
        _str(c.guardian_address_vn),                            # 27
        _str(c.guardian_address_jp),                            # 28
        _str(c.guardian_phone),                                 # 29
        _edu_period(edus[0]),                                   # 30
        _str(edus[0].school_name_vn if edus[0] else ""),        # 31
        _edu_period(edus[1]),                                   # 32
        _str(edus[1].school_name_vn if edus[1] else ""),        # 33
        _edu_period(edus[2]),                                   # 34
        _str(edus[2].school_name_vn if edus[2] else ""),        # 35
        _work_period(works[0]),                                 # 36
        _work_name(works[0]),                                   # 37
        _work_period(works[1]),                                 # 38
        _work_name(works[1]),                                   # 39
        _work_period(works[2]),                                 # 40
        _work_name(works[2]),                                   # 41
        _str(asgn.internship_field_vn if asgn else ""),         # 42
        _str(asgn.internship_field_jp if asgn else ""),         # 43
        _str(c.skill_summary_jp),                               # 44
        _str(c.skill_summary_vn),                               # 45
        _str(sup_org.name_vn if sup_org else ""),               # 46
        _str(sup_org.name_jp if sup_org else ""),               # 47
        _str(sup_org.address_vn if sup_org else ""),            # 48
        _str(sup_org.address_jp if sup_org else ""),            # 49
        _str(sup_org.representative_vn if sup_org else ""),     # 50
        _str(sup_org.representative_jp if sup_org else ""),     # 51
        _str(acc_org.name_vn if acc_org else ""),               # 52
        _str(acc_org.name_jp if acc_org else ""),               # 53
        _str(acc_org.address_vn if acc_org else ""),            # 54
        _str(acc_org.address_jp if acc_org else ""),            # 55
        _str(acc_org.representative_vn if acc_org else ""),     # 56
        _str(acc_org.representative_jp if acc_org else ""),     # 57
        _str(send_org.name_vn if send_org else ""),             # 58
        _str(send_org.representative_vn if send_org else ""),   # 59
        _str(c.phone),                                          # 60
    ]


def _row_to_candidate_dict(row: list) -> dict:
    """Chuyển 1 dòng Excel (60 giá trị) → dict candidate + relations."""
    def g(i):  # 0-indexed
        return _str(row[i]) if i < len(row) else ""

    return {
        "candidate": {
            "profile_code":       g(1),
            "full_name_vn":       g(2),
            "full_name_eng":      g(3),
            "full_name_katakana": g(4),
            "gender":             g(5),
            "date_of_birth":      _parse_date_str(row[16]) if len(row) > 16 else "",
            "date_of_birth_jp":   g(17),
            "address_vn":         g(20),
            "address_jp":         g(21),
            "birthplace_vn":      g(22),
            "birthplace_jp":      g(23),
            "guardian_name_vn":   g(24),
            "guardian_name_jp":   g(25),
            "guardian_address_vn":g(26),
            "guardian_address_jp":g(27),
            "guardian_phone":     g(28),
            "skill_summary_jp":   g(43),
            "skill_summary_vn":   g(44),
            "phone":              g(59),
            "status":             "draft",
        },
        "cccd": {
            "document_type":   "CCCD",
            "document_number": g(6),
            "issue_date":      _parse_date_str(row[7]) if len(row) > 7 else "",
            "issue_date_jp":   g(8),
            "issue_place_vn":  g(9),
            "issue_place_jp":  g(10),
        },
        "passport": {
            "document_type":   "Passport",
            "document_number": g(11),
            "issue_date":      _parse_date_str(row[12]) if len(row) > 12 else "",
            "issue_date_jp":   g(13),
            "issue_place_vn":  g(14),
            "issue_place_jp":  g(15),
        },
        "educations": [
            {"school_name_vn": g(30), "start_date": "", "end_date": "", "school_name_jp": "", "education_level": ""},
            {"school_name_vn": g(32), "start_date": "", "end_date": "", "school_name_jp": "", "education_level": ""},
            {"school_name_vn": g(34), "start_date": "", "end_date": "", "school_name_jp": "", "education_level": ""},
        ],
        "works": [
            {"company_name_vn": g(36), "start_date": "", "end_date": "", "company_name_jp": "", "job_title_vn": "", "job_title_jp": ""},
            {"company_name_vn": g(38), "start_date": "", "end_date": "", "company_name_jp": "", "job_title_vn": "", "job_title_jp": ""},
            {"company_name_vn": g(40), "start_date": "", "end_date": "", "company_name_jp": "", "job_title_vn": "", "job_title_jp": ""},
        ],
        "internship_field_vn": g(41),
        "internship_field_jp": g(42),
        "supervising_org": {
            "name_vn": g(45), "name_jp": g(46),
            "address_vn": g(47), "address_jp": g(48),
            "representative_vn": g(49), "representative_jp": g(50),
        },
        "accepting_org": {
            "name_vn": g(51), "name_jp": g(52),
            "address_vn": g(53), "address_jp": g(54),
            "representative_vn": g(55), "representative_jp": g(56),
        },
        "sending_org": {
            "name_vn": g(57), "representative_vn": g(58),
        },
    }


# ─────────────────────────────────────────────────────────────
# ENDPOINTS
# ─────────────────────────────────────────────────────────────

@excel_io_bp.route("/excel/config", methods=["GET"])
def get_excel_config():
    db = get_session()
    try:
        path = _get_excel_path(db)
        exists = os.path.isfile(path)
        return jsonify({"path": path, "exists": exists})
    finally:
        db.close()


@excel_io_bp.route("/excel/config", methods=["POST"])
def set_excel_config():
    db = get_session()
    try:
        data = request.get_json() or {}
        new_path = data.get("path", "").strip()
        if not new_path:
            return jsonify({"error": "path is required"}), 400
        row = db.query(AppSettings).filter(AppSettings.key == "excel_output_path").first()
        if row:
            row.value = new_path
        else:
            db.add(AppSettings(key="excel_output_path", value=new_path,
                               description="Đường dẫn file Excel lưu dữ liệu"))
        db.commit()
        return jsonify({"ok": True, "path": new_path})
    except Exception as e:
        db.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        db.close()


@excel_io_bp.route("/excel/preview", methods=["GET"])
def preview_excel():
    """Đọc Excel, trả về JSON preview (không import vào DB)."""
    db = get_session()
    try:
        path = _get_excel_path(db)
        if not os.path.isfile(path):
            return jsonify({"error": f"File không tồn tại: {path}"}), 404
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        rows_out = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(c for c in row):
                continue
            rows_out.append([_str(c) for c in row])
        return jsonify({"path": path, "headers": HEADERS, "rows": rows_out[:50]})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        db.close()


@excel_io_bp.route("/excel/import", methods=["POST"])
def import_from_excel():
    """Đọc File_lưu.xlsx → import vào SQLite (thêm mới, không ghi đè)."""
    db = get_session()
    try:
        data = request.get_json() or {}
        path = data.get("path") or _get_excel_path(db)
        if not os.path.isfile(path):
            return jsonify({"error": f"File không tồn tại: {path}"}), 404

        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active

        created = 0
        skipped = 0
        errors = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(c for c in row):
                continue
            row = list(row)
            try:
                d = _row_to_candidate_dict(row)
                cand_data = d["candidate"]
                name_vn = cand_data.get("full_name_vn", "").strip()
                if not name_vn:
                    skipped += 1
                    continue

                # Kiểm tra trùng theo profile_code hoặc tên + ngày sinh
                profile_code = cand_data.get("profile_code", "").strip()
                existing = None
                if profile_code:
                    existing = db.query(Candidate).filter(
                        Candidate.profile_code == profile_code).first()
                if not existing:
                    existing = db.query(Candidate).filter(
                        Candidate.full_name_vn == name_vn,
                        Candidate.date_of_birth == cand_data.get("date_of_birth", "")
                    ).first()

                if existing:
                    skipped += 1
                    continue

                # Tạo Candidate
                valid_cols = set(Candidate.__table__.columns.keys()) - {"id", "created_at"}
                c = Candidate(**{k: v for k, v in cand_data.items() if k in valid_cols})
                db.add(c)
                db.flush()

                # CCCD
                cccd = d["cccd"]
                if cccd.get("document_number"):
                    db.add(IdentityDocument(candidate_id=c.id, **cccd))

                # Passport
                psp = d["passport"]
                if psp.get("document_number"):
                    db.add(IdentityDocument(candidate_id=c.id, **psp))

                # Học vấn
                for edu in d["educations"]:
                    if edu.get("school_name_vn"):
                        valid = {k: v for k, v in edu.items()
                                 if k in Education.__table__.columns.keys() and k != "id"}
                        db.add(Education(candidate_id=c.id, **valid))

                # Công việc
                for w in d["works"]:
                    if w.get("company_name_vn"):
                        valid = {k: v for k, v in w.items()
                                 if k in WorkExperience.__table__.columns.keys() and k != "id"}
                        db.add(WorkExperience(candidate_id=c.id, **valid))

                # Nghiệp đoàn / Tiếp nhận / Phái cử
                def _get_or_create_org(org_data: dict, org_type: str):
                    name = org_data.get("name_vn", "").strip()
                    if not name:
                        return None
                    org = db.query(Organization).filter(
                        Organization.name_vn == name,
                        Organization.type == org_type
                    ).first()
                    if not org:
                        org = Organization(type=org_type, **{k: v for k, v in org_data.items()
                                                             if k in Organization.__table__.columns.keys()
                                                             and k not in ("id", "type")})
                        db.add(org)
                        db.flush()
                    return org

                sup = _get_or_create_org(d["supervising_org"], "supervising")
                acc = _get_or_create_org(d["accepting_org"], "accepting")
                send_data = d["sending_org"]
                send_org = None
                if send_data.get("name_vn"):
                    send_org = _get_or_create_org(send_data, "sending")

                asgn = CandidateAssignment(
                    candidate_id=c.id,
                    supervising_org_id=sup.id if sup else None,
                    accepting_org_id=acc.id if acc else None,
                    sending_org_id=send_org.id if send_org else None,
                    internship_field_vn=d.get("internship_field_vn", ""),
                    internship_field_jp=d.get("internship_field_jp", ""),
                )
                db.add(asgn)
                created += 1

            except Exception as e:
                errors.append({"row": row_idx, "error": str(e)})

        db.commit()
        return jsonify({
            "ok": True,
            "created": created,
            "skipped": skipped,
            "errors": errors,
        })
    except Exception as e:
        db.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        db.close()


@excel_io_bp.route("/excel/export", methods=["GET"])
def export_to_excel():
    """Xuất toàn bộ SQLite → ghi vào File_lưu.xlsx."""
    db = get_session()
    try:
        path = _get_excel_path(db)
        candidates = db.query(Candidate).order_by(Candidate.id).all()

        # Tạo workbook mới hoặc mở file hiện có để giữ sheet phụ
        if os.path.isfile(path):
            wb = openpyxl.load_workbook(path)
            # Xoá sheet chính cũ để ghi lại
            main_sheet_name = wb.sheetnames[0]
            del wb[main_sheet_name]
            ws = wb.create_sheet(main_sheet_name, 0)
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Data"

        # Header row
        header_fill = PatternFill("solid", fgColor="1A1A1A")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        for col_idx, h in enumerate(HEADERS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[cell.column_letter].width = 20

        # Data rows
        for stt, c in enumerate(candidates, start=1):
            row_data = _candidate_to_row(c, stt)
            for col_idx, val in enumerate(row_data, start=1):
                ws.cell(row=stt + 1, column=col_idx, value=val)

        wb.save(path)
        return jsonify({
            "ok": True,
            "path": path,
            "exported": len(candidates),
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        db.close()

@excel_io_bp.route("/export/template", methods=["GET"])
def export_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Form Mau"

    # Start with base HEADERS
    headers = list(HEADERS)

    # Add custom fields if any
    db = get_session()
    try:
        settings = db.query(AppSettings).filter(AppSettings.key == "custom_field_defs").first()
        if settings and settings.value:
            cfs = json.loads(settings.value)
            for f in cfs:
                headers.append(f.get("label", ""))
                if f.get("requireJp"):
                    headers.append(f.get("label", "") + " (JPN)")
    except Exception:
        pass
    finally:
        db.close()

    header_fill = PatternFill("solid", fgColor="1A1A1A")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20

    out = BytesIO()
    wb.save(out)
    out.seek(0)

    return send_file(
        out,
        as_attachment=True,
        download_name="Form_Mau_TTS.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
