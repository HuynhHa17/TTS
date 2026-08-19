"""API: Export to File_lưu.xlsx và PDF 履歴書"""
import os
import json
from io import BytesIO
import openpyxl
from flask import Blueprint, jsonify, send_file, request
from core.database import get_session
from core.models import Candidate, AppSettings, Organization, to_dict
from core.exporter import export_to_excel, _style_header_row, COL60_HEADERS
from core.pdf_exporter import build_rirekisho_pdf
from api.candidates import _build_full_profile
import config

export_bp = Blueprint("export", __name__)

@export_bp.route("/export/template", methods=["GET"])
def export_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Form Mau"

    headers = ["STT"] + COL60_HEADERS[1:]

    db = get_session()
    try:
        settings = db.query(AppSettings).first()
        if settings and settings.custom_field_defs:
            cfs = json.loads(settings.custom_field_defs)
            for f in cfs:
                headers.append(f.get("label", ""))
                if f.get("requireJp"):
                    headers.append(f.get("label", "") + " (JPN)")
    except Exception:
        pass
    finally:
        db.close()

    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=h)

    _style_header_row(ws, 1, len(headers))
    for c in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 20

    out = BytesIO()
    wb.save(out)
    out.seek(0)

    return send_file(
        out,
        as_attachment=True,
        download_name="Form_Mau_TTS.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _org_to_syndicate_dict(org) -> dict:
    """Map Organization (type=supervising) → dict cho build_syndicate_sheet."""
    return {
        "id":            org.id,
        "ten_vnm":       org.name_vn or "",
        "ten_jpn":       org.name_jp or "",
        "chu_tich_vnm":  org.representative_vn or "",
        "chu_tich_jpn":  org.representative_jp or "",
        "dia_chi_vnm":   org.address_vn or "",
        "dia_chi_jpn":   org.address_jp or "",
        "so_dien_thoai": org.phone or "",
    }


def _org_to_company_dict(org) -> dict:
    """Map Organization (type=accepting) → dict cho build_company_sheet."""
    return {
        "id":            org.id,
        "ten_vnm":       org.name_vn or "",
        "ten_jpn":       org.name_jp or "",
        "giam_doc_vnm":  org.representative_vn or "",
        "giam_doc_jpn":  org.representative_jp or "",
        "dia_chi_vnm":   org.address_vn or "",
        "dia_chi_jpn":   org.address_jp or "",
        "so_dien_thoai": org.phone or "",
    }


@export_bp.route("/export", methods=["GET", "POST"])
def export_excel():
    db = get_session()
    try:
        candidates = db.query(Candidate).order_by(Candidate.id).all()
        syndicates = db.query(Organization).filter(Organization.type == "supervising").order_by(Organization.id).all()
        companies  = db.query(Organization).filter(Organization.type == "accepting").order_by(Organization.id).all()

        c_dicts  = [to_dict(c) for c in candidates]
        s_dicts  = [_org_to_syndicate_dict(s) for s in syndicates]
        co_dicts = [_org_to_company_dict(co) for co in companies]

        row = db.query(AppSettings).filter(AppSettings.key == "excel_output_path").first()
        out_path = row.value if row and row.value else config.OUTPUT_FILE
        out_path = os.path.abspath(os.path.normpath(out_path))
    finally:
        db.close()

    sheet_name = request.args.get("sheet", None)
    target_send = out_path
    try:
        export_to_excel(c_dicts, s_dicts, co_dicts, out_path, sheet_name)
    except PermissionError:
        import tempfile
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        target_send = tmp.name
        tmp.close()
        export_to_excel(c_dicts, s_dicts, co_dicts, target_send, sheet_name)

    return send_file(
        target_send,
        as_attachment=True,
        download_name=os.path.basename(out_path) or "File_lưu.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@export_bp.route("/export/path", methods=["GET"])
def export_info():
    """Return output file path info."""
    db = get_session()
    try:
        row_out = db.query(AppSettings).filter(AppSettings.key == "excel_output_path").first()
        row_cv = db.query(AppSettings).filter(AppSettings.key == "template_cv_path").first()
        out_path = row_out.value if row_out and row_out.value else config.OUTPUT_FILE
        cv_path = row_cv.value if row_cv and row_cv.value else config.CV_FILE

        out_path = os.path.abspath(os.path.normpath(out_path))
        cv_path = os.path.abspath(os.path.normpath(cv_path))

        return jsonify({
            "output_path": out_path,
            "cv_path":     cv_path,
            "exists":      os.path.exists(out_path),
        })
    finally:
        db.close()


@export_bp.route("/export/pdf/<int:cid>", methods=["GET"])
def export_pdf_single(cid):
    """Xuất hồ sơ 1 ứng viên ra PDF dạng 履歴書."""
    db = get_session()
    try:
        c = db.query(Candidate).filter(Candidate.id == cid).first()
        if not c:
            return jsonify({"error": "Not found"}), 404
        profile = _build_full_profile(c)
    finally:
        db.close()

    try:
        pdf_bytes = build_rirekisho_pdf(profile)
    except Exception as e:
        return jsonify({"error": f"PDF generation failed: {e}"}), 500

    cand = profile.get("candidate", {})
    name = (cand.get("full_name_eng") or cand.get("full_name_vn") or f"candidate_{cid}").upper()
    code = cand.get("profile_code") or str(cid)
    # Lấy số thứ tự từ profile_code nếu có, ví dụ TTS-001 → 001
    stt = code.split("-")[-1] if "-" in str(code) else str(code)
    filename = f"{stt}. {name} - TCMMXD.pdf"

    return send_file(
        BytesIO(pdf_bytes),
        as_attachment=True,
        download_name=filename,
        mimetype="application/pdf",
    )


@export_bp.route("/export/pdf/all", methods=["GET"])
def export_pdf_all():
    """Xuất tất cả ứng viên, mỗi người 1 file PDF đóng gói vào ZIP."""
    import zipfile
    db = get_session()
    try:
        candidates = db.query(Candidate).order_by(Candidate.id).all()
        profiles   = [_build_full_profile(c) for c in candidates]
    finally:
        db.close()

    zip_buf = BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for idx, profile in enumerate(profiles, 1):
            try:
                pdf_bytes = build_rirekisho_pdf(profile)
            except Exception:
                continue
            cand = profile.get("candidate", {})
            name = (cand.get("full_name_eng") or cand.get("full_name_vn") or f"candidate_{idx}").upper()
            code = cand.get("profile_code") or str(idx)
            stt  = code.split("-")[-1] if "-" in str(code) else str(code)
            filename = f"{stt}. {name} - TCMMXD.pdf"
            zf.writestr(filename, pdf_bytes)

    zip_buf.seek(0)
    return send_file(
        zip_buf,
        as_attachment=True,
        download_name="TTS_HoSo_PDF.zip",
        mimetype="application/zip",
    )
