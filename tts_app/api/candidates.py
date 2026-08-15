import json
import os
from flask import Blueprint, request, jsonify, current_app
from sqlalchemy import or_
from core.database import get_session
from core.models import (
    Candidate, IdentityDocument, Education, WorkExperience,
    SkillExperience, JapanExperience, FamilyMember, CandidateAssignment, Organization,
    AppSettings, to_dict
)
import config

candidates_bp = Blueprint("candidates", __name__)

def _build_full_profile(c):
    c_dict = to_dict(c)
    if c_dict and c_dict.get("custom_fields"):
        try:
            c_dict["custom_fields"] = json.loads(c_dict["custom_fields"])
        except:
            c_dict["custom_fields"] = {}
            
    return {
        "candidate": c_dict,
        "identityDocuments": [to_dict(x) for x in c.identity_documents],
        "educations": [to_dict(x) for x in c.educations],
        "workExperiences": [to_dict(x) for x in c.work_experiences],
        "skillExperiences": [to_dict(x) for x in c.skill_experiences],
        "japanExperiences": [to_dict(x) for x in c.japan_experiences],
        "familyMembers": [to_dict(x) for x in c.family_members],
        "assignment": to_dict(c.assignment) if c.assignment else None,
        "supervisingOrg": to_dict(c.assignment.supervising_org) if c.assignment and c.assignment.supervising_org else None,
        "acceptingOrg": to_dict(c.assignment.accepting_org) if c.assignment and c.assignment.accepting_org else None,
        "sendingOrg": to_dict(c.assignment.sending_org) if c.assignment and c.assignment.sending_org else None,
    }


def _sync_excel():
    """Tự động xuất toàn bộ SQLite → File Excel sau mỗi thao tác thay đổi dữ liệu."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from datetime import datetime, date

        db = get_session()
        try:
            row = db.query(AppSettings).filter(AppSettings.key == "excel_output_path").first()
            path = row.value if row and row.value else config.OUTPUT_FILE

            candidates = db.query(Candidate).order_by(Candidate.id).all()

            HEADERS = [
                "STT", "MA HO SO", "TEN VNM", "TEN ENG", "TEN PHIEN AM", "GIOI TINH JPN",
                "SO CAN CUOC", "NGAY CAP CAN CUOC VNM", "NGAY CAP CAN CUOC JPN",
                "NOI CAP CAN CUOC VNM", "NOI CAP CAN CUOC JPN",
                "SO HO CHIEU", "NGAY CAP HO CHIEU VNM", "NGAY CAP HO CHIEU JPN",
                "NOI CAP HO CHIEU VNM", "NOI CAP HO CHIEU JPN",
                "NAM SINH VNM", "NAM SINH JPN",
                "TINH TRANG HON NHAN", "CO CON",
                "DIA CHI VNM", "DIA CHI JPN", "NOI SINH VNM", "NOI SINH JPN",
                "NGUOI GIAM HO VNM", "NGUOI GIAM HO JPN",
                "DIA CHI NGUOI GIAM HO VNM", "DIA CHI NGUOI GIAM HO JPN",
                "SDT NGUOI GIAM HO",
                "TRUONG HOC 1 VNM", "TRUONG HOC 1 JPN",
                "TRUONG HOC 2 VNM", "TRUONG HOC 2 JPN",
                "TRUONG HOC 3 VNM", "TRUONG HOC 3 JPN",
                "CONG TY 1 VNM", "CONG TY 1 JPN",
                "CONG TY 2 VNM", "CONG TY 2 JPN",
                "CONG TY 3 VNM", "CONG TY 3 JPN",
                "LINH VUC THUC TAP VNM", "LINH VUC THUC TAP JPN",
                "TOM TAT KN JPN", "TOM TAT KN VNM",
                "NGHIEP DOAN VNM", "NGHIEP DOAN JPN",
                "DC NGHIEP DOAN VNM", "DC NGHIEP DOAN JPN",
                "ND NGHIEP DOAN VNM", "ND NGHIEP DOAN JPN",
                "CONG TY TIEP NHAN VNM", "CONG TY TIEP NHAN JPN",
                "DC CONG TY TIEP NHAN VNM", "DC CONG TY TIEP NHAN JPN",
                "ND CONG TY TIEP NHAN VNM", "ND CONG TY TIEP NHAN JPN",
                "CONG TY PHAI CU VNM", "ND CONG TY PHAI CU VNM",
                "SO DIEN THOAI",
            ]

            def _str(v):
                if v is None: return ""
                if isinstance(v, (datetime, date)): return str(v)
                return str(v).strip()

            def _fmt_vn(val):
                if not val: return ""
                s = str(val).strip()
                for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d", "%d-%m-%Y"):
                    try:
                        d = datetime.strptime(s, fmt)
                        return d.strftime("%d/%m/%Y")
                    except ValueError:
                        pass
                return s

            def _fmt_jp(val):
                if not val: return ""
                s = str(val).strip()
                if "年" in s and "月" in s: return s
                for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d", "%d-%m-%Y"):
                    try:
                        d = datetime.strptime(s, fmt)
                        return f"{d.year}年{d.month:02d}月{d.day:02d}日"
                    except ValueError:
                        pass
                return s

            def _get_doc(c, dtype):
                for d in c.identity_documents:
                    if d.document_type == dtype:
                        return d
                return None

            def _get_edu(c, idx):
                edus = c.educations
                return edus[idx] if idx < len(edus) else None

            def _get_work(c, idx):
                works = c.work_experiences
                return works[idx] if idx < len(works) else None

            def _get_org(c, org_type):
                if c.assignment:
                    if org_type == "supervising" and c.assignment.supervising_org:
                        return c.assignment.supervising_org
                    if org_type == "accepting" and c.assignment.accepting_org:
                        return c.assignment.accepting_org
                    if org_type == "sending" and c.assignment.sending_org:
                        return c.assignment.sending_org
                return None

            # Tạo/mở workbook
            if os.path.isfile(path):
                wb = openpyxl.load_workbook(path)
                main_sheet_name = wb.sheetnames[0]
                del wb[main_sheet_name]
                ws = wb.create_sheet(main_sheet_name, 0)
            else:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Data"

            # Header
            header_fill = PatternFill("solid", fgColor="1A1A1A")
            header_font = Font(bold=True, color="FFFFFF", size=10)
            for col_idx, h in enumerate(HEADERS, start=1):
                cell = ws.cell(row=1, column=col_idx, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws.column_dimensions[cell.column_letter].width = 18

            # Data rows
            for stt, c in enumerate(candidates, start=1):
                cccd = _get_doc(c, "CCCD")
                psp  = _get_doc(c, "Passport")
                edu0 = _get_edu(c, 0)
                edu1 = _get_edu(c, 1)
                edu2 = _get_edu(c, 2)
                wrk0 = _get_work(c, 0)
                wrk1 = _get_work(c, 1)
                wrk2 = _get_work(c, 2)
                sup_org = _get_org(c, "supervising")
                acc_org = _get_org(c, "accepting")
                snd_org = _get_org(c, "sending")

                row_data = [
                    stt,
                    _str(c.profile_code),
                    _str(c.full_name_vn),
                    _str(c.full_name_eng),
                    _str(c.full_name_katakana),
                    _str(c.gender),
                    _str(cccd.document_number if cccd else ""),
                    _fmt_vn(cccd.issue_date if cccd else ""),
                    _fmt_jp(cccd.issue_date_jp if cccd else ""),
                    _str(cccd.issue_place_vn if cccd else ""),
                    _str(cccd.issue_place_jp if cccd else ""),
                    _str(psp.document_number if psp else ""),
                    _fmt_vn(psp.issue_date if psp else ""),
                    _fmt_jp(psp.issue_date_jp if psp else ""),
                    _str(psp.issue_place_vn if psp else ""),
                    _str(psp.issue_place_jp if psp else ""),
                    _fmt_vn(c.date_of_birth),
                    _fmt_jp(c.date_of_birth_jp or c.date_of_birth),
                    _str(c.marital_status),
                    _str(c.has_children),
                    _str(c.address_vn),
                    _str(c.address_jp),
                    _str(c.birthplace_vn),
                    _str(c.birthplace_jp),
                    _str(c.guardian_name_vn),
                    _str(c.guardian_name_jp),
                    _str(c.guardian_address_vn),
                    _str(c.guardian_address_jp),
                    _str(c.guardian_phone),
                    _str(edu0.school_name_vn if edu0 else ""),
                    _str(edu0.school_name_jp if edu0 else ""),
                    _str(edu1.school_name_vn if edu1 else ""),
                    _str(edu1.school_name_jp if edu1 else ""),
                    _str(edu2.school_name_vn if edu2 else ""),
                    _str(edu2.school_name_jp if edu2 else ""),
                    _str(wrk0.company_name_vn if wrk0 else ""),
                    _str(wrk0.company_name_jp if wrk0 else ""),
                    _str(wrk1.company_name_vn if wrk1 else ""),
                    _str(wrk1.company_name_jp if wrk1 else ""),
                    _str(wrk2.company_name_vn if wrk2 else ""),
                    _str(wrk2.company_name_jp if wrk2 else ""),
                    _str(c.skill_summary_vn),
                    _str(c.skill_summary_jp),
                    _str(c.skill_summary_jp),
                    _str(c.skill_summary_vn),
                    _str(sup_org.name_vn if sup_org else ""),
                    _str(sup_org.name_jp if sup_org else ""),
                    _str(sup_org.address_vn if sup_org else ""),
                    _str(sup_org.address_jp if sup_org else ""),
                    _str(sup_org.representative_vn if sup_org else ""),
                    _str(sup_org.representative_jp if sup_org else ""),
                    _str(acc_org.name_vn if acc_org else ""),
                    _str(acc_org.name_jp if acc_org else ""),
                    _str(acc_org.address_vn if acc_org else ""),
                    _str(acc_org.address_jp if acc_org else ""),
                    _str(acc_org.representative_vn if acc_org else ""),
                    _str(acc_org.representative_jp if acc_org else ""),
                    _str(snd_org.name_vn if snd_org else ""),
                    _str(snd_org.representative_vn if snd_org else ""),
                    _str(c.phone),
                ]

                for col_idx, val in enumerate(row_data, start=1):
                    ws.cell(row=stt + 1, column=col_idx, value=val)

            wb.save(path)
            current_app.logger.info(f"[AutoSync] Đã ghi {len(candidates)} hồ sơ → {path}")
        finally:
            db.close()
    except Exception as e:
        # Không làm hỏng response nếu sync excel thất bại
        try:
            current_app.logger.warning(f"[AutoSync] Lỗi khi ghi Excel: {e}")
        except:
            pass


@candidates_bp.route("/candidates", methods=["GET"])
def list_candidates():
    db = get_session()
    try:
        q      = request.args.get("q", "").strip()
        page   = int(request.args.get("page", 1))
        limit  = int(request.args.get("limit", 50))
        status = request.args.get("status", "")

        query = db.query(Candidate)
        if q:
            query = query.filter(or_(
                Candidate.full_name_vn.ilike(f"%{q}%"),
                Candidate.profile_code.ilike(f"%{q}%"),
                Candidate.phone.ilike(f"%{q}%"),
            ))
        if status:
            query = query.filter(Candidate.status == status)

        total  = query.count()
        items  = query.order_by(Candidate.id).offset((page - 1) * limit).limit(limit).all()

        results = []
        for c in items:
            c_dict = to_dict(c)
            if c_dict and c_dict.get("custom_fields"):
                try:
                    c_dict["custom_fields"] = json.loads(c_dict["custom_fields"])
                except:
                    c_dict["custom_fields"] = {}
            results.append(c_dict)

        return jsonify(results)
    finally:
        db.close()

@candidates_bp.route("/candidates/<int:cid>", methods=["GET"])
def get_candidate(cid):
    db = get_session()
    try:
        c = db.query(Candidate).filter(Candidate.id == cid).first()
        if not c:
            return jsonify({"error": "Not found"}), 404
        return jsonify(_build_full_profile(c))
    finally:
        db.close()

@candidates_bp.route("/candidates", methods=["POST"])
def create_candidate():
    db = get_session()
    try:
        data = request.get_json() or {}
        
        # Sanitize empty strings to None for SQLAlchemy
        def _sanitize(d):
            if isinstance(d, dict):
                return {k: (None if v == "" else v) for k, v in d.items()}
            return d

        # Skip child records where all meaningful fields are empty
        def _is_empty(d):
            skip = {'id', 'candidate_id'}
            return all(v is None for k, v in d.items() if k not in skip)

        c_data = _sanitize(data.get("candidate", {}))
        
        count = db.query(Candidate).count()
        if not c_data.get("profile_code"):
            c_data["profile_code"] = f"TTS-{count + 1:03d}"
            
        if "custom_fields" in c_data and isinstance(c_data["custom_fields"], dict):
            c_data["custom_fields"] = json.dumps(c_data["custom_fields"])
            
        valid_cols = set(Candidate.__table__.columns.keys()) - {"id", "created_at", "updated_at"}
        c = Candidate(**{k: v for k, v in c_data.items() if k in valid_cols})
        db.add(c)
        db.flush()

        for doc in data.get("identityDocuments", []):
            doc = _sanitize(doc)
            if not _is_empty(doc):
                db.add(IdentityDocument(**{k: v for k, v in doc.items() if k in IdentityDocument.__table__.columns.keys() and k not in ('id', 'candidate_id')}, candidate_id=c.id))
        
        for edu in data.get("educations", []):
            edu = _sanitize(edu)
            if not _is_empty(edu):
                db.add(Education(**{k: v for k, v in edu.items() if k in Education.__table__.columns.keys() and k not in ('id', 'candidate_id')}, candidate_id=c.id))

        for work in data.get("workExperiences", []):
            work = _sanitize(work)
            if not _is_empty(work):
                db.add(WorkExperience(**{k: v for k, v in work.items() if k in WorkExperience.__table__.columns.keys() and k not in ('id', 'candidate_id')}, candidate_id=c.id))
            
        for skill in data.get("skillExperiences", []):
            skill = _sanitize(skill)
            if not _is_empty(skill):
                db.add(SkillExperience(**{k: v for k, v in skill.items() if k in SkillExperience.__table__.columns.keys() and k not in ('id', 'candidate_id')}, candidate_id=c.id))
            
        for jpe in data.get("japanExperiences", []):
            jpe = _sanitize(jpe)
            if not _is_empty(jpe):
                db.add(JapanExperience(**{k: v for k, v in jpe.items() if k in JapanExperience.__table__.columns.keys() and k not in ('id', 'candidate_id')}, candidate_id=c.id))
            
        for fam in data.get("familyMembers", []):
            fam = _sanitize(fam)
            if not _is_empty(fam):
                db.add(FamilyMember(**{k: v for k, v in fam.items() if k in FamilyMember.__table__.columns.keys() and k not in ('id', 'candidate_id')}, candidate_id=c.id))

        db.commit()
        db.refresh(c)
        result = _build_full_profile(c)
        _sync_excel()   # ← Tự động cập nhật Excel
        return jsonify(result), 201
    except Exception as e:
        db.rollback()
        return jsonify({"error": str(e)}), 400
    finally:
        db.close()

@candidates_bp.route("/candidates/<int:cid>", methods=["PUT"])
def update_candidate(cid):
    db = get_session()
    try:
        c = db.query(Candidate).filter(Candidate.id == cid).first()
        if not c:
            return jsonify({"error": "Not found"}), 404
            
        data = request.get_json() or {}
        
        def _sanitize(d):
            if isinstance(d, dict):
                return {k: (None if v == "" else v) for k, v in d.items()}
            return d

        def _is_empty(d):
            skip = {'id', 'candidate_id'}
            return all(v is None for k, v in d.items() if k not in skip)

        c_data = _sanitize(data.get("candidate", {}))
        
        valid_cols = set(Candidate.__table__.columns.keys()) - {"id", "created_at", "updated_at"}
        
        if "custom_fields" in c_data and isinstance(c_data["custom_fields"], dict):
            c_data["custom_fields"] = json.dumps(c_data["custom_fields"])
            
        for k, v in c_data.items():
            if k in valid_cols:
                setattr(c, k, v)
                
        # We will keep it simple: drop all related records and recreate
        # In a production app you'd want to merge them by ID
        for rel in [c.identity_documents, c.educations, c.work_experiences, 
                    c.skill_experiences, c.japan_experiences, c.family_members]:
            for item in list(rel):
                db.delete(item)
                
        db.flush()
        
        for doc in data.get("identityDocuments", []):
            doc = _sanitize(doc)
            if not _is_empty(doc):
                db.add(IdentityDocument(**{k: v for k, v in doc.items() if k in IdentityDocument.__table__.columns.keys() and k not in ('id', 'candidate_id')}, candidate_id=c.id))
        for edu in data.get("educations", []):
            edu = _sanitize(edu)
            if not _is_empty(edu):
                db.add(Education(**{k: v for k, v in edu.items() if k in Education.__table__.columns.keys() and k not in ('id', 'candidate_id')}, candidate_id=c.id))
        for work in data.get("workExperiences", []):
            work = _sanitize(work)
            if not _is_empty(work):
                db.add(WorkExperience(**{k: v for k, v in work.items() if k in WorkExperience.__table__.columns.keys() and k not in ('id', 'candidate_id')}, candidate_id=c.id))
        for skill in data.get("skillExperiences", []):
            skill = _sanitize(skill)
            if not _is_empty(skill):
                db.add(SkillExperience(**{k: v for k, v in skill.items() if k in SkillExperience.__table__.columns.keys() and k not in ('id', 'candidate_id')}, candidate_id=c.id))
        for jpe in data.get("japanExperiences", []):
            jpe = _sanitize(jpe)
            if not _is_empty(jpe):
                db.add(JapanExperience(**{k: v for k, v in jpe.items() if k in JapanExperience.__table__.columns.keys() and k not in ('id', 'candidate_id')}, candidate_id=c.id))
        for fam in data.get("familyMembers", []):
            fam = _sanitize(fam)
            if not _is_empty(fam):
                db.add(FamilyMember(**{k: v for k, v in fam.items() if k in FamilyMember.__table__.columns.keys() and k not in ('id', 'candidate_id')}, candidate_id=c.id))

        db.commit()
        db.refresh(c)
        result = _build_full_profile(c)
        _sync_excel()   # ← Tự động cập nhật Excel
        return jsonify(result)
    except Exception as e:
        db.rollback()
        return jsonify({"error": str(e)}), 400
    finally:
        db.close()

@candidates_bp.route("/candidates/<int:cid>", methods=["DELETE"])
def delete_candidate(cid):
    db = get_session()
    try:
        c = db.query(Candidate).filter(Candidate.id == cid).first()
        if not c:
            return jsonify({"error": "Not found"}), 404
        db.delete(c)
        db.commit()
        _sync_excel()   # ← Tự động cập nhật Excel
        return jsonify({"ok": True})
    except Exception as e:
        db.rollback()
        return jsonify({"error": str(e)}), 400
    finally:
        db.close()

@candidates_bp.route("/candidates/batch-delete", methods=["POST"])
def batch_delete_candidates():
    db = get_session()
    try:
        data = request.get_json() or {}
        ids = data.get("ids", [])
        if not ids:
            return jsonify({"error": "Danh sách ID trống"}), 400
        candidates = db.query(Candidate).filter(Candidate.id.in_(ids)).all()
        for c in candidates:
            db.delete(c)
        db.commit()
        _sync_excel()   # ← Tự động cập nhật Excel
        return jsonify({"ok": True, "deleted": len(candidates)})
    except Exception as e:
        db.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        db.close()
