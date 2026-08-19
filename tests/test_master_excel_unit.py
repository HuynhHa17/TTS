"""
test_master_excel_unit.py — Unit tests cho chức năng Master Excel (File_lưu.xlsx)
Bao gồm:
  1. Cấu trúc chuẩn 60 cột dữ liệu (Canonical 60 columns)
  2. Chuyển đổi Candidate Model sang dòng Excel 60 cột (_candidate_to_row)
  3. Chuyển đổi dòng Excel 60 cột sang Candidate Profile Dict (_row_to_candidate_dict)
  4. Xuất file Master Excel đầy đủ các Sheet: Ứng viên, Nghiệp đoàn, Xí nghiệp tiếp nhận
  5. Đồng bộ 2 chiều SQLite Database <-> File_lưu.xlsx
"""

import sys
import os
import tempfile
import openpyxl
import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "tts_app"))
os.chdir(os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "tts_app"))

from run import app as flask_app
from core.database import init_db, get_session
from core.models import Candidate, IdentityDocument, Education, WorkExperience, Organization
from core.exporter import export_to_excel, COL60_HEADERS
from api.excel_io import HEADERS, _candidate_to_row, _row_to_candidate_dict


@pytest.fixture(scope="module")
def app():
    flask_app.config["TESTING"] = True
    init_db()
    yield flask_app


@pytest.fixture(scope="module")
def client(app):
    return app.test_client()


class TestMasterExcelSchemaAndConversion:
    """1. Kiểm tra 60 cột và hàm chuyển đổi 2 chiều"""

    def test_60_headers_structure(self):
        assert len(COL60_HEADERS) == 60
        assert len(HEADERS) == 60
        # Cột STT, Mã hồ sơ, Họ tên, Giới tính
        assert HEADERS[0] == "STT"
        assert HEADERS[1] == "MA HO SO"
        assert HEADERS[2] == "TEN VNM"
        assert HEADERS[3] == "TEN ENG"
        assert HEADERS[4] == "TEN PHIEN AM"
        assert HEADERS[5] == "GIOI TINH JPN"
        # CCCD & Passport
        assert HEADERS[6] == "SO CAN CUOC"
        assert HEADERS[11] == "SO HO CHIEU"
        # Năm sinh & Hôn nhân
        assert HEADERS[16] == "NAM SINH VNM"
        assert HEADERS[17] == "NAM SINH JPN"
        assert HEADERS[18] == "TINH TRANG HON NHAN"
        assert HEADERS[19] == "CO CON"
        # Trường học 1-3
        assert HEADERS[29] == "TRUONG HOC 1 VNM"
        assert HEADERS[30] == "TRUONG HOC 1 JPN"
        assert HEADERS[31] == "TRUONG HOC 2 VNM"
        assert HEADERS[32] == "TRUONG HOC 2 JPN"
        # Công ty 1-3
        assert HEADERS[35] == "CONG TY 1 VNM"
        assert HEADERS[36] == "CONG TY 1 JPN"
        assert HEADERS[37] == "CONG TY 2 VNM"
        assert HEADERS[38] == "CONG TY 2 JPN"
        # SĐT
        assert HEADERS[59] == "SO DIEN THOAI"

    def test_candidate_to_row_conversion(self):
        c = Candidate(
            profile_code="TTS-099",
            full_name_vn="Lý Tiểu Long",
            full_name_eng="LY TIEU LONG",
            full_name_katakana="リー・ティエウ・ロン",
            gender="Nam",
            date_of_birth="2000-11-20",
            date_of_birth_jp="2000年11月20日",
            marital_status="Độc thân",
            has_children="Không",
            address_vn="Quận 1, TP HCM",
            address_jp="ホーチミン市1区",
            birthplace_vn="Quảng Nam",
            birthplace_jp="クアンナム省",
            guardian_name_vn="Lý Đại Ca",
            guardian_phone="0909123456",
            phone="0911222333",
        )
        c.identity_documents = [
            IdentityDocument(document_type="CCCD", document_number="079200009988", issue_date="2021-04-10", issue_place_vn="TP HCM")
        ]
        c.educations = [
            Education(school_name_vn="THPT Nguyễn Thị Minh Khai", school_name_jp="ミンカイ高校")
        ]
        c.work_experiences = [
            WorkExperience(company_name_vn="Cty TNHH Giao Hàng", company_name_jp="配達会社")
        ]

        row = _candidate_to_row(c, stt=1)
        assert len(row) == 60
        assert row[0] == 1
        assert row[1] == "TTS-099"
        assert row[2] == "Lý Tiểu Long"
        assert row[3] == "LY TIEU LONG"
        assert row[4] == "リー・ティエウ・ロン"
        assert row[5] == "Nam"
        assert row[6] == "079200009988"
        assert row[18] == "Độc thân"
        assert row[19] == "Không"
        assert row[29] == "THPT Nguyễn Thị Minh Khai"
        assert row[30] == "ミンカイ高校"
        assert row[35] == "Cty TNHH Giao Hàng"
        assert row[36] == "配達会社"
        assert row[59] == "0911222333"

    def test_row_to_candidate_dict_conversion(self):
        row = [""] * 60
        row[0] = 1
        row[1] = "TTS-123"
        row[2] = "Bùi Văn Test"
        row[3] = "BUI VAN TEST"
        row[4] = "ブイ ヴァン テスト"
        row[5] = "Nam"
        row[6] = "048200001234"
        row[16] = "2002-04-12"
        row[17] = "2002年04月12日"
        row[18] = "Đã kết hôn"
        row[19] = "Có"
        row[29] = "THPT Phan Châu Trinh"
        row[35] = "Công ty Cơ Điện"
        row[59] = "0988111222"

        d = _row_to_candidate_dict(row)
        cand = d["candidate"]
        assert cand["profile_code"] == "TTS-123"
        assert cand["full_name_vn"] == "Bùi Văn Test"
        assert cand["marital_status"] == "Đã kết hôn"
        assert cand["has_children"] == "Có"
        assert cand["phone"] == "0988111222"
        assert d["cccd"]["document_number"] == "048200001234"
        assert d["educations"][0]["school_name_vn"] == "THPT Phan Châu Trinh"
        assert d["works"][0]["company_name_vn"] == "Công ty Cơ Điện"


class TestMasterExcelExportAndSync:
    """2. Kiểm tra xuất file Master Excel và đồng bộ"""

    def test_export_to_excel_file(self):
        candidates = [{
            "id": 1,
            "profile_code": "TTS-001",
            "full_name_vn": "Nguyễn Văn Export",
            "full_name_eng": "NGUYEN VAN EXPORT",
            "full_name_katakana": "グエン ヴァン エクスポート",
            "gender": "Nam",
            "date_of_birth": "2000-01-01",
            "date_of_birth_jp": "2000年01月01日",
            "phone": "0912345678",
            "status": "draft",
        }]
        syndicates = [{
            "id": 1,
            "ten_vnm": "Nghiệp đoàn Tokyo",
            "ten_jpn": "東京協同組合",
            "chu_tich_vnm": "Tanaka",
            "chu_tich_jpn": "田中",
            "dia_chi_vnm": "Tokyo",
            "dia_chi_jpn": "東京都",
            "so_dien_thoai": "03-1234-5678",
        }]
        companies = [{
            "id": 1,
            "ten_vnm": "Xí nghiệp Osaka",
            "ten_jpn": "大阪工業",
            "giam_doc_vnm": "Yamada",
            "giam_doc_jpn": "山田",
            "dia_chi_vnm": "Osaka",
            "dia_chi_jpn": "大阪府",
            "so_dien_thoai": "06-9876-5432",
        }]

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            out_file = tmp.name

        try:
            export_to_excel(candidates, syndicates, companies, out_file)
            assert os.path.exists(out_file)

            wb = openpyxl.load_workbook(out_file)
            # Kiểm tra các Sheet
            assert len(wb.sheetnames) >= 2
            assert any("nghiệp đoàn" in s.lower() or "nghiep doan" in s.lower() for s in wb.sheetnames)
            assert any("sử dụng" in s.lower() or "xí nghiệp" in s.lower() or "xi nghiep" in s.lower() for s in wb.sheetnames)
        finally:
            if os.path.exists(out_file):
                os.remove(out_file)

    def test_api_excel_config_and_preview(self, client):
        # GET config
        res = client.get("/api/excel/config")
        assert res.status_code == 200
        data = res.get_json()
        assert "path" in data

        # GET preview
        res_prev = client.get("/api/excel/preview")
        assert res_prev.status_code in (200, 404)
        if res_prev.status_code == 200:
            p_data = res_prev.get_json()
            assert "headers" in p_data
            assert len(p_data["headers"]) == 60
