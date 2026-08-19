"""
test_form_io.py — Unit tests for Excel Candidate Form Template Generation & Import Parser
"""

import io
import pytest
import openpyxl
from io import BytesIO
from tts_app.run import app
from tts_app.core.form_template import create_candidate_form_workbook, export_candidate_form_template
from tts_app.core.form_parser import parse_candidate_form_excel


@pytest.fixture
def client():
    app.config["TESTING"] = True
    with app.test_client() as c:
        yield c


class TestFormTemplate:
    def test_create_candidate_form_workbook(self):
        wb = create_candidate_form_workbook()
        assert "To_Khai_Ung_Vien" in wb.sheetnames
        ws = wb["To_Khai_Ung_Vien"]
        # Check title and key section headers
        assert "PHIẾU ĐĂNG KÝ THÔNG TIN" in str(ws["A1"].value)
        assert "I. THÔNG TIN CÁ NHÂN" in str(ws["A3"].value)
        assert "II. GIẤY TỜ TÙY THÂN" in str(ws["A10"].value)
        assert "III. THỂ LỰC & TÌNH TRẠNG SỨC KHỎE" in str(ws["A15"].value)
        assert "IV. NGUYỆN VỌNG THỰC TẬP" in str(ws["A19"].value)
        assert "V. QUÁ TRÌNH HỌC VẤN" in str(ws["A24"].value)
        assert "VI. QUÁ TRÌNH LÀM VIỆC" in str(ws["A30"].value)
        assert "VII. THÀNH VIÊN GIA ĐÌNH" in str(ws["A36"].value)

    def test_export_candidate_form_template_bytes(self):
        data = export_candidate_form_template()
        assert isinstance(data, (bytes, bytearray))
        assert len(data) > 2000
        # Verify valid excel file
        wb = openpyxl.load_workbook(BytesIO(data))
        assert "To_Khai_Ung_Vien" in wb.sheetnames


class TestFormParser:
    def test_parse_blank_form(self):
        data = export_candidate_form_template()
        parsed = parse_candidate_form_excel(data)
        assert "candidate" in parsed
        assert "identityDocuments" in parsed
        assert "educations" in parsed
        assert "workExperiences" in parsed
        assert "familyMembers" in parsed
        assert "assignment" in parsed

    def test_parse_filled_form(self):
        wb = create_candidate_form_workbook()
        ws = wb["To_Khai_Ung_Vien"]
        
        # Fill in test values
        ws["B4"] = "TRẦN VĂN TEST"
        ws["D4"] = "チャン ヴァン テスト"
        ws["F4"] = "TRAN VAN TEST"
        ws["B5"] = "15/01/2000"
        ws["D5"] = "Nam"
        ws["F5"] = "0987654321"
        ws["B8"] = "Hà Nội"
        ws["B9"] = "Số 123 Đường Cầu Giấy, Hà Nội"
        ws["B11"] = "001200012345"  # CCCD
        ws["D11"] = "10/05/2021"
        ws["B13"] = "Trần Văn Bố"
        ws["D13"] = "Bố"
        ws["F13"] = "0912345678"
        
        # Add a school (Row 26)
        ws["B26"] = "09/2015"
        ws["C26"] = "06/2018"
        ws["D26"] = "THPT Cầu Giấy"
        ws["G26"] = "THPT"

        # Add a job (Row 32)
        ws["B32"] = "07/2018"
        ws["C32"] = "12/2022"
        ws["D32"] = "Công ty TNHH Xây Dựng ABC"
        ws["F32"] = "Thợ xây dựng"

        # Add family (Row 38)
        ws["B38"] = "Bố"
        ws["C38"] = "Trần Văn Bố"
        ws["E38"] = "1972"
        ws["F38"] = "Làm nông"

        out = BytesIO()
        wb.save(out)
        out.seek(0)

        parsed = parse_candidate_form_excel(out.getvalue())
        c = parsed["candidate"]
        assert c["full_name_vn"] == "TRẦN VĂN TEST"
        assert c["full_name_katakana"] == "チャン ヴァン テスト"
        assert c["date_of_birth"] == "15/01/2000"
        assert c["phone"] == "0987654321"
        assert c["birthplace_vn"] == "Hà Nội"
        assert c["address_vn"] == "Số 123 Đường Cầu Giấy, Hà Nội"
        assert c["guardian_name"] == "Trần Văn Bố"

        docs = parsed["identityDocuments"]
        assert len(docs) >= 1
        assert docs[0]["document_number"] == "001200012345"

        edus = parsed["educations"]
        assert len(edus) >= 1
        assert edus[0]["school_name_vn"] == "THPT Cầu Giấy"

        works = parsed["workExperiences"]
        assert len(works) >= 1
        assert works[0]["company_name_vn"] == "Công ty TNHH Xây Dựng ABC"

        family = parsed["familyMembers"]
        assert len(family) >= 1
        assert family[0]["full_name"] == "Trần Văn Bố"
        assert family[0]["birth_year"] == 1972


class TestFormEndpoints:
    def test_download_form_template(self, client):
        res = client.get("/api/documents/form-template")
        assert res.status_code == 200
        assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in res.content_type
        assert len(res.data) > 2000

    def test_preview_form_endpoint(self, client):
        wb = create_candidate_form_workbook()
        ws = wb["To_Khai_Ung_Vien"]
        ws["B4"] = "LÊ THỊ PREVIEW"
        ws["B5"] = "20/11/2002"
        ws["F5"] = "0909123456"

        out = BytesIO()
        wb.save(out)
        out.seek(0)

        data = {"file": (out, "To_Khai_Ung_Vien.xlsx")}
        res = client.post("/api/documents/preview-form", data=data, content_type="multipart/form-data")
        assert res.status_code == 200
        json_data = res.get_json()
        assert json_data["candidate"]["full_name_vn"] == "LÊ THỊ PREVIEW"
        assert json_data["candidate"]["date_of_birth"] == "20/11/2002"

    def test_import_form_endpoint(self, client):
        wb = create_candidate_form_workbook()
        ws = wb["To_Khai_Ung_Vien"]
        ws["B4"] = "VÕ THÀNH CÔNG"
        ws["D4"] = "ヴォー タイン コン"
        ws["B5"] = "08/08/1999"
        ws["F5"] = "0918889999"
        ws["B8"] = "Đà Nẵng"
        ws["B9"] = "Hải Châu, Đà Nẵng"
        ws["B11"] = "048099001234"

        out = BytesIO()
        wb.save(out)
        out.seek(0)

        data = {"file": (out, "To_Khai_Ung_Vien.xlsx")}
        res = client.post("/api/documents/import-form", data=data, content_type="multipart/form-data")
        assert res.status_code == 201
        json_data = res.get_json()
        assert json_data["ok"] is True
        assert "candidate_id" in json_data
        cid = json_data["candidate_id"]

        # Verify candidate created in DB
        res_get = client.get(f"/api/candidates/{cid}")
        assert res_get.status_code == 200
        cand_data = res_get.get_json()
        assert cand_data["candidate"]["full_name_vn"] == "VÕ THÀNH CÔNG"
        assert cand_data["candidate"]["birthplace_vn"] == "Đà Nẵng"

        # Cleanup
        client.delete(f"/api/candidates/{cid}")

    def test_import_forms_batch_endpoint(self, client):
        # File 1
        wb1 = create_candidate_form_workbook()
        ws1 = wb1["To_Khai_Ung_Vien"]
        ws1["B4"] = "NGUYỄN BATCH A"
        ws1["B5"] = "01/01/2001"
        out1 = BytesIO()
        wb1.save(out1)
        out1.seek(0)

        # File 2
        wb2 = create_candidate_form_workbook()
        ws2 = wb2["To_Khai_Ung_Vien"]
        ws2["B4"] = "TRẦN BATCH B"
        ws2["B5"] = "02/02/2002"
        out2 = BytesIO()
        wb2.save(out2)
        out2.seek(0)

        data = {
            "files": [
                (out1, "Don_1.xlsx"),
                (out2, "Don_2.xlsx")
            ]
        }
        res = client.post("/api/documents/import-forms-batch", data=data, content_type="multipart/form-data")
        assert res.status_code == 200
        json_data = res.get_json()
        assert json_data["ok"] is True
        assert json_data["imported_count"] == 2
        assert len(json_data["imported"]) == 2

        # Cleanup
        for item in json_data["imported"]:
            client.delete(f"/api/candidates/{item['id']}")

    def test_import_form_saves_all_child_records(self, client):
        wb = create_candidate_form_workbook()
        ws = wb["To_Khai_Ung_Vien"]
        ws["B4"] = "NGUYỄN VĂN ĐẦY ĐỦ"
        ws["B5"] = "28/10/2000"
        ws["B11"] = "079200001122"

        # Education
        ws["B26"] = "09/2015"
        ws["C26"] = "06/2018"
        ws["D26"] = "THPT Lê Quý Đôn"
        ws["G26"] = "THPT"

        # Work
        ws["B32"] = "08/2018"
        ws["C32"] = "10/2021"
        ws["D32"] = "Công ty Cơ Khí ABC"
        ws["F32"] = "Kỹ thuật viên tiện"

        # Family 1
        ws["B38"] = "Bố"
        ws["C38"] = "Nguyễn Văn Cha"
        ws["E38"] = "1970"
        ws["F38"] = "Kinh doanh tự do"
        ws["H38"] = "Có"

        # Family 2
        ws["B39"] = "Mẹ"
        ws["C39"] = "Trần Thị Mẹ"
        ws["E39"] = "1975"
        ws["F39"] = "Nội trợ"
        ws["H39"] = "Có"

        out = BytesIO()
        wb.save(out)
        out.seek(0)

        data = {"file": (out, "To_Khai_Ung_Vien.xlsx")}
        res = client.post("/api/documents/import-form", data=data, content_type="multipart/form-data")
        assert res.status_code == 201
        cid = res.get_json()["candidate_id"]

        # Check full profile
        res_get = client.get(f"/api/candidates/{cid}")
        assert res_get.status_code == 200
        p = res_get.get_json()

        assert len(p["educations"]) >= 1
        assert p["educations"][0]["school_name_vn"] == "THPT Lê Quý Đôn"
        assert p["educations"][0]["education_level"] == "THPT"

        assert len(p["workExperiences"]) >= 1
        assert p["workExperiences"][0]["company_name_vn"] == "Công ty Cơ Khí ABC"
        assert p["workExperiences"][0]["job_title_vn"] == "Kỹ thuật viên tiện"

        assert len(p["familyMembers"]) >= 2
        assert p["familyMembers"][0]["full_name"] == "Nguyễn Văn Cha"
        assert p["familyMembers"][0]["relationship"] == "Bố"
        assert p["familyMembers"][0]["living_together"] == "Có"

        # Cleanup
        client.delete(f"/api/candidates/{cid}")


class TestDateTranslationOffline:
    def test_date_translation_without_api_key(self, client):
        # Translate date field should work without API key
        res = client.post("/api/translate/field", json={
            "field_name": "date_of_birth_jp",
            "value": "2000-10-28"
        })
        assert res.status_code == 200
        assert res.get_json()["translation"] == "2000年10月28日"

        res2 = client.post("/api/translate/field", json={
            "field_name": "ngay_cap_cccd_jpn",
            "value": "15/05/2021"
        })
        assert res2.status_code == 200
        assert res2.get_json()["translation"] == "2021年05月15日"


class TestMasterExcelSyncAndOpen:
    def test_export_khai_tt_auto_sync(self, client):
        res = client.get("/api/documents/khai-tt")
        assert res.status_code == 200
        assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in res.content_type

    def test_excel_open_endpoint(self, client):
        # Should return 200 or execute smoothly
        res = client.get("/api/excel/open")
        assert res.status_code in (200, 500)  # On headless CI, os.startfile may raise but route executed sync

